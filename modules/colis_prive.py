"""
Module Colis Privé pour GREENLOG
Croisement fichiers logisticien et Colis Privé avec détection majorations
Version 1.0 - Intégration architecture modulaire
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime
from shared import persistence

def create_excel_with_format(df):
    """Créer un fichier Excel formaté avec mise en forme"""
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Croisement', engine='openpyxl')
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb['Croisement']
    
    # Couleurs GREENLOG
    vert = PatternFill(start_color='6BBFA3', end_color='6BBFA3', fill_type='solid')
    bleu = PatternFill(start_color='2D3E50', end_color='2D3E50', fill_type='solid')
    rouge = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # En-tête en bleu marine
    for cell in ws[1]:
        cell.fill = bleu
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Lignes avec majorations en rouge
    for row in range(2, ws.max_row + 1):
        majoration_cell = ws.cell(row, 9)  # Colonne Majoration service
        if majoration_cell.value and majoration_cell.value > 0:
            for col in range(1, 10):
                ws.cell(row, col).fill = rouge
    
    # Ajuster largeurs colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)
    
    return excel_output

def run():
    """Point d'entrée principal du module Colis Privé"""
    
    st.title("🚚 Module Colis Privé")
    st.markdown("Croisement fichiers logisticien et Colis Privé avec détection des majorations")
    st.markdown("---")
    
    # Initialisation session_state spécifique au module
    if 'colis_prive_files' not in st.session_state:
        st.session_state.colis_prive_files = {}
    if 'colis_prive_data' not in st.session_state:
        st.session_state.colis_prive_data = None
    if 'colis_prive_auto_loaded' not in st.session_state:
        st.session_state.colis_prive_auto_loaded = False
    
    # 💾 CHARGEMENT AUTOMATIQUE au premier démarrage du module
    if not st.session_state.colis_prive_auto_loaded:
        saved_files = persistence.load_module_files('colis_prive')
        saved_data = persistence.load_module_data('colis_prive')
        
        if saved_files:
            st.session_state.colis_prive_files = saved_files
            st.info(f"✅ {len(saved_files)} fichier(s) Colis Privé chargé(s) depuis la sauvegarde")
        
        if saved_data:
            st.session_state.colis_prive_data = saved_data
        
        st.session_state.colis_prive_auto_loaded = True
    
    # Sidebar pour les uploads et contrôles
    with st.sidebar:
        st.header("📁 Fichiers")
        
        # Fichiers logisticien depuis bibliothèque
        st.subheader("📋 Fichiers Logisticien")
        
        from modules.logisticiens_library import load_logisticien_files_for_analysis, get_all_available_periods
        
        available_files = load_logisticien_files_for_analysis(nb_months=3)
        
        if len(available_files) > 0:
            st.success(f"✅ {len(available_files)} fichier(s) chargé(s) automatiquement depuis la bibliothèque")
            periods = get_all_available_periods()[:len(available_files)]
            for period in periods:
                from modules.bibliotheque import get_month_name
                st.caption(f"• {get_month_name(period['month'])} {period['year']}")
        else:
            st.warning("⚠️ Aucun fichier logisticien dans la bibliothèque")
            st.caption("Allez dans le module 📋 Logisticiens pour en ajouter")
        
        st.markdown("---")
        
        # Upload fichier Colis Privé
        st.subheader("Fichier Colis Privé")
        fichier_cp = st.file_uploader(
            "Fichier CSV Colis Privé",
            type=['csv'],
            key="colis_prive_csv",
            help="Uploadez le fichier CSV Colis Privé"
        )
        
        if fichier_cp:
            st.session_state.colis_prive_files['csv'] = fichier_cp
            st.success(f"✅ {fichier_cp.name}")
        elif 'csv' in st.session_state.colis_prive_files:
            st.info(f"📄 {st.session_state.colis_prive_files['csv'].name}")
        
        st.markdown("---")
        
        # Bouton d'analyse
        can_analyze = (fichier_cp or 'csv' in st.session_state.colis_prive_files) and len(available_files) > 0
        
        if not can_analyze:
            if len(available_files) == 0:
                st.warning("⚠️ Ajoutez des fichiers logisticiens dans le module 📋 Logisticiens")
            else:
                st.info("💡 Uploadez un fichier CSV Colis Privé pour lancer l'analyse")
        
        if st.button("🚀 Lancer l'analyse", type="primary", disabled=not can_analyze, use_container_width=True):
            with st.spinner("Analyse en cours..."):
                try:
                    # Charger le fichier Colis Privé
                    fichier_to_process = fichier_cp if fichier_cp else st.session_state.colis_prive_files.get('csv')
                    fichier_to_process.seek(0)
                    df_cp = pd.read_csv(fichier_to_process, sep=';', encoding='utf-8-sig', decimal=',')
                    
                    # Charger les fichiers logisticien depuis la bibliothèque
                    all_log_data = []
                    for log_file in available_files:
                        try:
                            log_file.seek(0)
                            df_log_temp = pd.read_excel(log_file, sheet_name='Facturation préparation')
                            all_log_data.append(df_log_temp)
                        except Exception as e:
                            st.warning(f"⚠️ Erreur lecture {log_file.name}: {str(e)}")
                    
                    if not all_log_data:
                        st.error("❌ Aucun fichier logisticien valide")
                    else:
                        df_log = pd.concat(all_log_data, ignore_index=True)
                        df_log = df_log.drop_duplicates(subset=['Numéro de tracking'], keep='first')
                        
                        # Sélectionner et renommer colonnes
                        df_log = df_log[[
                            'Nom du partenaire',
                            "Numéro de commande d'origine",
                            'Numéro de commande partenaire',
                            'Numéro de tracking',
                            'Date de la commande',
                            'Poids expédition'
                        ]].copy()
                        
                        df_cp = df_cp[[
                            'Tracking',
                            'Poids facturé',
                            'Majoration service',
                            'Code Postal'
                        ]].copy()
                        
                        # Nettoyer les données
                        df_log['Poids expédition'] = pd.to_numeric(df_log['Poids expédition'], errors='coerce')
                        
                        for col in ['Poids facturé', 'Majoration service']:
                            if df_cp[col].dtype == 'object':
                                df_cp[col] = df_cp[col].astype(str).str.replace(',', '.').astype(float)
                            else:
                                df_cp[col] = pd.to_numeric(df_cp[col], errors='coerce')
                        
                        # Merger
                        df = pd.merge(df_log, df_cp, left_on='Numéro de tracking', right_on='Tracking', how='inner')
                        df = df.drop('Tracking', axis=1)
                        
                        # Réorganiser colonnes
                        df = df[[
                            'Nom du partenaire',
                            "Numéro de commande d'origine",
                            'Numéro de commande partenaire',
                            'Numéro de tracking',
                            'Date de la commande',
                            'Poids expédition',
                            'Poids facturé',
                            'Code Postal',
                            'Majoration service'
                        ]]
                        
                        # Trier par majoration décroissante
                        df = df.sort_values('Majoration service', ascending=False)
                        
                        # Sauvegarder les données
                        st.session_state.colis_prive_data = {
                            'df': df,
                            'timestamp': datetime.now()
                        }
                        
                        # 💾 SAUVEGARDE AUTOMATIQUE
                        persistence.save_module_files('colis_prive', st.session_state.colis_prive_files)
                        persistence.save_module_data('colis_prive', st.session_state.colis_prive_data)
                        
                        # 📚 AUTO-ARCHIVAGE DANS LA BIBLIOTHÈQUE
                        success, year, month = persistence.auto_archive_analysis(
                            'Colis_Prive',
                            df,
                            st.session_state.colis_prive_data
                        )
                        
                        # Message avec info archivage
                        if success:
                            from modules.bibliotheque import get_month_name
                            st.success(f"✅ Analyse terminée et archivée ({get_month_name(month)} {year})")
                        else:
                            st.success("✅ Analyse terminée et sauvegardée !")
                        
                        st.rerun()
                
                except Exception as e:
                    st.error(f"❌ Erreur: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
        
        # Bouton réinitialisation
        st.markdown("---")
        if st.button("🗑️ Réinitialiser", type="secondary", use_container_width=True):
            persistence.delete_module_data('colis_prive')
            st.session_state.colis_prive_files = {}
            st.session_state.colis_prive_data = None
            st.session_state.colis_prive_auto_loaded = False
            st.success("✅ Module réinitialisé !")
            st.rerun()
    
    # Affichage principal
    if st.session_state.colis_prive_data is None:
        st.info("👈 Uploadez vos fichiers et lancez l'analyse")
        
        st.markdown("""
        ### 📋 Fichiers requis:
        1. **Fichiers Logisticien** - Uploadés depuis la page d'accueil (partagés entre modules)
        2. **Fichier Colis Privé** (CSV) - À uploader dans ce module
        
        ### ✨ Fonctionnalités:
        - 🔄 Croisement automatique des fichiers
        - 🔍 Détection des majorations de service
        - 📊 Statistiques détaillées
        - 📥 Export Excel formaté (majorations en rouge)
        - 💾 Sauvegarde automatique
        """)
    
    else:
        data = st.session_state.colis_prive_data
        df = data['df']
        
        # Statistiques globales
        col1, col2, col3, col4 = st.columns(4)
        
        nb_majorations = len(df[df['Majoration service'] > 0])
        total_majorations = df['Majoration service'].sum()
        
        with col1:
            st.metric("Total lignes croisées", len(df))
        with col2:
            st.metric("Lignes avec majorations", nb_majorations)
        with col3:
            st.metric("Total majorations", f"{total_majorations:.2f} €", delta_color="inverse")
        with col4:
            pct_majorations = (nb_majorations / len(df) * 100) if len(df) > 0 else 0
            st.metric("% avec majorations", f"{pct_majorations:.1f}%")
        
        st.markdown("---")
        
        # Tabs
        tab1, tab2, tab3 = st.tabs([
            "📊 Vue d'ensemble",
            "⚠️ Majorations détectées",
            "📋 Détail complet"
        ])
        
        # TAB 1: VUE D'ENSEMBLE
        with tab1:
            st.header("📊 Vue d'ensemble")
            
            # Top 10 des majorations
            st.subheader("🔝 Top 10 des majorations les plus élevées")
            df_top = df[df['Majoration service'] > 0].head(10)
            
            if len(df_top) > 0:
                df_display = df_top[[
                    'Numéro de tracking',
                    'Nom du partenaire',
                    'Code Postal',
                    'Majoration service'
                ]].copy()
                
                df_display['Majoration service'] = df_display['Majoration service'].apply(lambda x: f"{x:.2f} €")
                
                st.dataframe(df_display, use_container_width=True, hide_index=True)
            else:
                st.success("✅ Aucune majoration détectée !")
            
            # Répartition par partenaire
            st.subheader("👥 Répartition des majorations par partenaire")
            df_by_partner = df.groupby('Nom du partenaire').agg({
                'Numéro de tracking': 'count',
                'Majoration service': ['sum', lambda x: (x > 0).sum()]
            }).reset_index()
            
            df_by_partner.columns = ['Partenaire', 'Total envois', 'Total majorations (€)', 'Nb avec majoration']
            df_by_partner['Total majorations (€)'] = df_by_partner['Total majorations (€)'].round(2)
            df_by_partner = df_by_partner.sort_values('Total majorations (€)', ascending=False)
            
            st.dataframe(df_by_partner, use_container_width=True, hide_index=True)
        
        # TAB 2: MAJORATIONS
        with tab2:
            st.header("⚠️ Majorations détectées")
            
            df_maj = df[df['Majoration service'] > 0].copy()
            
            if len(df_maj) > 0:
                st.warning(f"⚠️ {len(df_maj)} ligne(s) avec majoration(s) détectée(s)")
                
                # Filtres
                col1, col2 = st.columns(2)
                
                with col1:
                    partenaires = ['Tous'] + sorted(df_maj['Nom du partenaire'].unique().tolist())
                    selected_partner = st.selectbox("Filtrer par partenaire:", partenaires, key="maj_partner")
                
                with col2:
                    min_maj = st.number_input("Majoration minimale (€):", min_value=0.0, value=0.0, step=0.1)
                
                # Filtrer
                if selected_partner != 'Tous':
                    df_maj = df_maj[df_maj['Nom du partenaire'] == selected_partner]
                
                if min_maj > 0:
                    df_maj = df_maj[df_maj['Majoration service'] >= min_maj]
                
                st.info(f"📊 {len(df_maj)} ligne(s) affichée(s) (après filtres)")
                
                # Afficher
                df_maj_display = df_maj[[
                    'Numéro de tracking',
                    'Nom du partenaire',
                    "Numéro de commande d'origine",
                    'Code Postal',
                    'Poids expédition',
                    'Poids facturé',
                    'Majoration service'
                ]].copy()
                
                df_maj_display['Majoration service'] = df_maj_display['Majoration service'].apply(lambda x: f"{x:.2f} €")
                
                st.dataframe(df_maj_display, use_container_width=True, hide_index=True)
                
                # Total après filtres
                total_filtered = df_maj['Majoration service'].sum()
                st.metric("💰 Total majorations (filtré)", f"{total_filtered:.2f} €")
            
            else:
                st.success("✅ Aucune majoration détectée dans ce croisement !")
        
        # TAB 3: DÉTAIL COMPLET
        with tab3:
            st.header("📋 Détail complet du croisement")
            
            col1, col2 = st.columns(2)
            
            with col1:
                partenaires_all = ['Tous'] + sorted(df['Nom du partenaire'].unique().tolist())
                selected_partner_all = st.selectbox("Filtrer par partenaire:", partenaires_all, key="all_partner")
            
            with col2:
                show_only_maj = st.checkbox("Afficher uniquement les majorations", value=False)
            
            df_filtered = df.copy()
            
            if selected_partner_all != 'Tous':
                df_filtered = df_filtered[df_filtered['Nom du partenaire'] == selected_partner_all]
            
            if show_only_maj:
                df_filtered = df_filtered[df_filtered['Majoration service'] > 0]
            
            st.info(f"📊 {len(df_filtered)} ligne(s) affichée(s)")
            
            # Afficher
            st.dataframe(df_filtered, use_container_width=True, hide_index=True)
        
        # Bouton d'export
        st.markdown("---")
        st.subheader("📥 Export Excel")
        
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            if st.button("📥 Télécharger Excel (avec majorations en rouge)", use_container_width=True):
                with st.spinner("Création du fichier Excel..."):
                    excel_data = create_excel_with_format(df)
                    
                    st.download_button(
                        label="✅ Télécharger Croisement_Colis_Prive.xlsx",
                        data=excel_data,
                        file_name=f"Croisement_Colis_Prive_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

if __name__ == "__main__":
    run()
