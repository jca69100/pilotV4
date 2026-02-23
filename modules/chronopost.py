"""
Module Chronopost pour GREENLOG EC
Analyse des factures Chronopost avec écarts et surplus
Version 2.1 - Intégration architecture modulaire
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from shared import persistence

# ============================================================================
# GRILLES TARIFAIRES INTÉGRÉES
# ============================================================================

TARIFS_FRANCE = {
    '0.00 - 0.50': 2.45, '0.50 - 1.00': 2.54, '1.00 - 2.00': 2.87,
    '2.00 - 3.00': 3.30, '3.00 - 4.00': 3.80, '4.00 - 5.00': 4.25,
    '5.00 - 6.00': 4.70, '6.00 - 7.00': 5.20, '7.00 - 8.00': 5.60,
    '8.00 - 9.00': 7.26, '9.00 - 10.00': 6.55, '10.00 - 11.00': 7.00,
    '11.00 - 12.00': 7.50, '12.00 - 13.00': 7.95, '13.00 - 14.00': 8.40,
    '14.00 - 15.00': 8.85, '15.00 - 16.00': 9.35, '16.00 - 17.00': 9.80,
    '17.00 - 18.00': 10.25, '18.00 - 19.00': 10.75, '19.00 - 20.00': 11.20
}

TARIFS_EUROPE = {
    '0.00 - 0.50': {'AT': 9.00, 'BE': 3.68, 'BG': 8.72, 'CH': 9.00, 'CZ': 8.72, 'DE': 4.45, 'DK': 8.72, 'EE': 8.72, 'ES': 6.01, 'FI': 8.72, 'HR': 8.72, 'HU': 8.72, 'IE': 8.72, 'IT': 6.01, 'LT': 8.72, 'LU': 4.45, 'LV': 8.72, 'NL': 4.45, 'PL': 8.72, 'PT': 8.72, 'RO': 8.72, 'SE': 9.89, 'SI': 9.00, 'SK': 9.00},
    '0.50 - 1.00': {'AT': 9.23, 'BE': 3.80, 'BG': 9.00, 'CH': 9.23, 'CZ': 9.00, 'DE': 4.75, 'DK': 9.00, 'EE': 9.00, 'ES': 6.36, 'FI': 9.00, 'HR': 9.00, 'HU': 9.00, 'IE': 9.00, 'IT': 6.36, 'LT': 9.00, 'LU': 4.75, 'LV': 9.00, 'NL': 4.75, 'PL': 9.00, 'PT': 9.00, 'RO': 9.00, 'SE': 10.12, 'SI': 9.23, 'SK': 9.23},
    '1.00 - 2.00': {'AT': 9.58, 'BE': 4.15, 'BG': 10.00, 'CH': 10.23, 'CZ': 10.00, 'DE': 5.50, 'DK': 10.00, 'EE': 10.00, 'ES': 8.97, 'FI': 10.00, 'HR': 10.00, 'HU': 10.00, 'IE': 10.00, 'IT': 8.97, 'LT': 10.00, 'LU': 5.50, 'LV': 10.00, 'NL': 5.50, 'PL': 10.00, 'PT': 10.00, 'RO': 10.00, 'SE': 10.47, 'SI': 9.58, 'SK': 9.58},
    '2.00 - 3.00': {'AT': 12.19, 'BE': 6.88, 'BG': 11.70, 'CH': 12.00, 'CZ': 11.70, 'DE': 7.36, 'DK': 11.70, 'EE': 11.70, 'ES': 9.66, 'FI': 11.70, 'HR': 11.70, 'HU': 11.70, 'IE': 11.70, 'IT': 9.66, 'LT': 11.70, 'LU': 8.05, 'LV': 11.70, 'NL': 7.36, 'PL': 11.70, 'PT': 11.70, 'RO': 11.70, 'SE': 13.08, 'SI': 12.19, 'SK': 12.19},
    '3.00 - 5.00': {'AT': 12.88, 'BE': 7.57, 'BG': 14.40, 'CH': 15.00, 'CZ': 14.40, 'DE': 8.74, 'DK': 14.40, 'EE': 14.40, 'ES': 10.35, 'FI': 14.40, 'HR': 14.40, 'HU': 14.40, 'IE': 14.40, 'IT': 10.35, 'LT': 14.40, 'LU': 9.43, 'LV': 14.40, 'NL': 8.74, 'PL': 14.40, 'PT': 14.40, 'RO': 14.40, 'SE': 13.77, 'SI': 12.88, 'SK': 12.88},
    '5.00 - 7.00': {'AT': 14.95, 'BE': 9.64, 'BG': 17.15, 'CH': 18.00, 'CZ': 17.15, 'DE': 10.12, 'DK': 17.15, 'EE': 17.15, 'ES': 11.73, 'FI': 17.15, 'HR': 17.15, 'HU': 17.15, 'IE': 17.15, 'IT': 11.73, 'LT': 17.15, 'LU': 11.50, 'LV': 17.15, 'NL': 10.12, 'PL': 17.15, 'PT': 17.15, 'RO': 17.15, 'SE': 15.84, 'SI': 14.95, 'SK': 14.95},
    '7.00 - 10.00': {'AT': 16.33, 'BE': 11.02, 'BG': 21.25, 'CH': 22.50, 'CZ': 21.25, 'DE': 11.50, 'DK': 21.25, 'EE': 21.25, 'ES': 13.11, 'FI': 21.25, 'HR': 21.25, 'HU': 21.25, 'IE': 21.25, 'IT': 13.11, 'LT': 21.25, 'LU': 13.88, 'LV': 21.25, 'NL': 11.50, 'PL': 21.25, 'PT': 21.25, 'RO': 21.25, 'SE': 17.22, 'SI': 16.33, 'SK': 16.33},
    '10.00 - 15.00': {'AT': 20.47, 'BE': 15.16, 'BG': 26.75, 'CH': 28.50, 'CZ': 26.75, 'DE': 16.33, 'DK': 26.75, 'EE': 26.75, 'ES': 17.94, 'FI': 26.75, 'HR': 26.75, 'HU': 26.75, 'IE': 26.75, 'IT': 17.94, 'LT': 26.75, 'LU': 20.16, 'LV': 26.75, 'NL': 16.33, 'PL': 26.75, 'PT': 26.75, 'RO': 26.75, 'SE': 21.36, 'SI': 20.47, 'SK': 20.47}
}

# ============================================================================
# FONCTIONS UTILITAIRES
# ============================================================================

def get_theoretical_price(weight_kg, country):
    """Calcul du prix théorique selon le poids et le pays"""
    if pd.isna(weight_kg) or weight_kg <= 0:
        return None
    
    # Déterminer la tranche de poids
    if weight_kg <= 0.5:
        weight_range = '0.00 - 0.50'
    elif weight_kg <= 1.0:
        weight_range = '0.50 - 1.00'
    elif weight_kg <= 2.0:
        weight_range = '1.00 - 2.00'
    elif weight_kg <= 3.0:
        weight_range = '2.00 - 3.00'
    elif weight_kg <= 4.0:
        weight_range = '3.00 - 4.00'
    elif weight_kg <= 5.0:
        weight_range = '4.00 - 5.00' if country == 'FR' else '3.00 - 5.00'
    elif weight_kg <= 6.0:
        weight_range = '5.00 - 6.00'
    elif weight_kg <= 7.0:
        weight_range = '6.00 - 7.00' if country == 'FR' else '5.00 - 7.00'
    elif weight_kg <= 8.0:
        weight_range = '7.00 - 8.00'
    elif weight_kg <= 9.0:
        weight_range = '8.00 - 9.00'
    elif weight_kg <= 10.0:
        weight_range = '9.00 - 10.00' if country == 'FR' else '7.00 - 10.00'
    elif weight_kg <= 11.0:
        weight_range = '10.00 - 11.00'
    elif weight_kg <= 12.0:
        weight_range = '11.00 - 12.00'
    elif weight_kg <= 13.0:
        weight_range = '12.00 - 13.00'
    elif weight_kg <= 14.0:
        weight_range = '13.00 - 14.00'
    elif weight_kg <= 15.0:
        weight_range = '14.00 - 15.00' if country == 'FR' else '10.00 - 15.00'
    elif weight_kg <= 16.0:
        weight_range = '15.00 - 16.00'
    elif weight_kg <= 17.0:
        weight_range = '16.00 - 17.00'
    elif weight_kg <= 18.0:
        weight_range = '17.00 - 18.00'
    elif weight_kg <= 19.0:
        weight_range = '18.00 - 19.00'
    elif weight_kg <= 20.0:
        weight_range = '19.00 - 20.00'
    else:
        return None
    
    # Chercher le prix
    if country == 'FR':
        return TARIFS_FRANCE.get(weight_range)
    else:
        if weight_range in TARIFS_EUROPE:
            return TARIFS_EUROPE[weight_range].get(country)
    
    return None

def load_chronopost_invoice(uploaded_file):
    """Charger et parser une facture Chronopost"""
    df_raw = pd.read_excel(uploaded_file, sheet_name='Table 1', header=None)
    
    header_lines = []
    for i in range(len(df_raw)):
        row = df_raw.iloc[i]
        has_date = False
        has_tracking = False
        
        for j in range(min(10, len(row))):
            if pd.notna(row[j]):
                val_str = str(row[j]).strip()
                if val_str in ['Date', 'DATE']:
                    has_date = True
                if 'objet' in val_str.lower():
                    has_tracking = True
        
        if has_date and has_tracking:
            header_lines.append(i)
    
    data_rows = []
    for section_idx, header_line in enumerate(header_lines):
        header_row = df_raw.iloc[header_line]
        
        date_col = None
        tracking_col = None
        poids_col = None
        montant_col = None
        obs_col = None
        
        for j, val in enumerate(header_row):
            if pd.notna(val):
                val_str = str(val).strip().lower()
                if val_str in ['date', 'DATE']:
                    date_col = j
                elif 'objet' in val_str:
                    tracking_col = j
                elif val_str == 'poids':
                    poids_col = j
                elif 'montant' in val_str:
                    montant_col = j
                elif val_str == 'obs':
                    obs_col = j
        
        next_header = header_lines[section_idx + 1] if section_idx + 1 < len(header_lines) else len(df_raw)
        
        for i in range(header_line + 3, next_header):
            row = df_raw.iloc[i]
            
            tracking = None
            if tracking_col is not None:
                for col_offset in [0, 1, 2, -1, -2]:
                    col_to_check = tracking_col + col_offset
                    if 0 <= col_to_check < len(row) and pd.notna(row[col_to_check]):
                        val = str(row[col_to_check]).strip()
                        if len(val) > 10 and any(val.startswith(p) for p in ['XR', 'XA', 'XT', '2L', '6A', 'LD', 'MH']):
                            tracking = val
                            break
            
            if tracking:
                date_value = None
                if date_col is not None and pd.notna(row[date_col]):
                    date_value = row[date_col]
                elif pd.notna(row[1]):
                    date_value = row[1]
                elif pd.notna(row[2]):
                    date_value = row[2]
                
                data_rows.append({
                    'Date': date_value,
                    'Tracking': tracking,
                    'Poids_Chronopost': row[poids_col] if poids_col and pd.notna(row[poids_col]) else None,
                    'Prix_Facture_HT': row[montant_col] if montant_col and pd.notna(row[montant_col]) else None,
                    'Observations': str(row[obs_col]) if obs_col and pd.notna(row[obs_col]) else ''
                })
    
    if data_rows:
        df = pd.DataFrame(data_rows)
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df = df[df['Date'].notna()]
        return df
    
    return pd.DataFrame()

def extract_surplus(df_raw):
    """Extraire les surplus d'une facture"""
    surplus_data = []
    current_tracking = None
    current_date = None
    
    for i in range(100, len(df_raw)):
        row = df_raw.iloc[i]
        
        for col in [3, 4, 5, 6]:
            if col < len(row) and pd.notna(row[col]):
                val = str(row[col]).strip()
                if len(val) > 10 and any(val.startswith(p) for p in ['XR', 'XT', 'XA', '2L', '6A', 'LD', 'MH']):
                    current_tracking = val
                    for date_col in [1, 2]:
                        if date_col < len(row) and pd.notna(row[date_col]):
                            try:
                                current_date = pd.to_datetime(row[date_col])
                                break
                            except:
                                pass
                    break
        
        surplus_type = None
        montant = None
        
        row_text = ' '.join([str(v) for v in row if pd.notna(v)]).upper()
        
        if 'ETIQUETTE' in row_text and 'NON CONFORME' in row_text:
            surplus_type = 'Etiquette non conforme'
        elif 'RETOUR' in row_text and 'EXPEDITEUR' in row_text:
            surplus_type = 'Retour expéditeur'
        elif 'TRAITEMENT' in row_text and 'RETOUR' in row_text:
            surplus_type = 'Traitement Retour expéditeur'
        elif 'ZONE' in row_text and ('DIFFICILE' in row_text or 'ELOIGNE' in row_text):
            surplus_type = 'Zones Difficiles d\'accès'
        elif 'CORSE' in row_text:
            surplus_type = 'Supplément Corse'
        elif 'HORS NORME' in row_text or 'HORS-NORME' in row_text:
            for col in range(len(row)):
                if pd.notna(row[col]) and isinstance(row[col], (int, float)):
                    if 60 < row[col] < 80:
                        surplus_type = 'Supplément hors norme'
                        break
                    elif 15 < row[col] < 25:
                        surplus_type = 'Supplément manutention'
                        break
            if not surplus_type:
                if 'MANUTENTION' in row_text:
                    surplus_type = 'Supplément manutention'
                else:
                    surplus_type = 'Supplément hors norme'
        elif 'MANUTENTION' in row_text:
            surplus_type = 'Supplément manutention'
        
        if surplus_type:
            for col in range(len(row)):
                if pd.notna(row[col]) and isinstance(row[col], (int, float)):
                    if 0 < row[col] < 100:
                        montant = row[col]
                        break
            
            if montant and current_tracking:
                surplus_data.append({
                    'Date': current_date,
                    'Tracking': current_tracking,
                    'Type_Surplus': surplus_type,
                    'Montant_Surplus': montant
                })
    
    return surplus_data

def create_excel_export(df_filtered, title):
    """Créer un export Excel formaté"""
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = title
    
    # Écrire les données
    for r_idx, row in enumerate([df_filtered.columns.tolist()] + df_filtered.values.tolist(), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # En-tête
            if r_idx == 1:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajuster largeurs
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    ws.freeze_panes = 'A2'
    wb.save(output)
    return output.getvalue()

# ============================================================================
# FONCTION PRINCIPALE DU MODULE
# ============================================================================

def run():
    """Point d'entrée principal du module Chronopost"""
    
    st.title("📦 Module Chronopost")
    st.markdown("Analyse des factures Chronopost avec détection des écarts et surplus")
    st.markdown("---")
    
    # Initialisation session_state spécifique au module
    if 'chronopost_files' not in st.session_state:
        st.session_state.chronopost_files = {}
    if 'chronopost_data' not in st.session_state:
        st.session_state.chronopost_data = None
    if 'chronopost_auto_loaded' not in st.session_state:
        st.session_state.chronopost_auto_loaded = False
    
    # 💾 CHARGEMENT AUTOMATIQUE au premier démarrage du module
    if not st.session_state.chronopost_auto_loaded:
        saved_files = persistence.load_module_files('chronopost')
        saved_data = persistence.load_module_data('chronopost')
        
        if saved_files:
            st.session_state.chronopost_files = saved_files
            st.info(f"✅ {len(saved_files)} fichier(s) Chronopost chargé(s) depuis la sauvegarde")
        
        if saved_data:
            st.session_state.chronopost_data = saved_data
        
        st.session_state.chronopost_auto_loaded = True
    
    # Sidebar pour les uploads et contrôles
    with st.sidebar:
        st.header("📁 Fichiers")
        
        # Upload factures Chronopost
        st.subheader("Factures Chronopost")
        factures = st.file_uploader(
            "Factures (.xlsx)",
            type=['xlsx'],
            accept_multiple_files=True,
            key="chronopost_factures",
            help="Uploadez une ou plusieurs factures Chronopost"
        )
        
        if factures:
            st.session_state.chronopost_files['factures'] = factures
            st.success(f"✅ {len(factures)} facture(s)")
        elif 'factures' in st.session_state.chronopost_files:
            st.info(f"📄 {len(st.session_state.chronopost_files['factures'])} facture(s) déjà chargée(s)")
        
        # Fichiers logisticien depuis bibliothèque
        st.subheader("📋 Fichiers Logisticien")
        
        from modules.logisticiens_library import load_logisticien_files_for_analysis, get_all_available_periods
        
        available_files = load_logisticien_files_for_analysis(nb_months=3)
        
        if len(available_files) > 0:
            st.success(f"✅ {len(available_files)} fichier(s) chargé(s) automatiquement depuis la bibliothèque")
            periods = get_all_available_periods()[:len(available_files)]
            for period in periods:
                from modules.bibliotheque import get_month_name
                st.caption(f"• {get_month_name(period['month'])} {period['year']}")
        else:
            st.warning("⚠️ Aucun fichier logisticien dans la bibliothèque")
            st.caption("Allez dans le module 📋 Logisticiens pour en ajouter")
        
        st.markdown("---")
        
        # Bouton d'analyse
        can_analyze = factures or 'factures' in st.session_state.chronopost_files
        
        if not can_analyze:
            st.info("💡 Uploadez au moins une facture Chronopost pour lancer l'analyse")
        
        if len(available_files) == 0:
            st.warning("⚠️ Sans fichiers logisticiens, seule l'analyse des factures et surplus sera disponible")
        
        if st.button("🚀 Lancer l'analyse", type="primary", disabled=not can_analyze, use_container_width=True):
            with st.spinner("Analyse en cours..."):
                try:
                    # Charger factures
                    all_invoices = []
                    all_surplus = []
                    
                    factures_to_process = factures if factures else st.session_state.chronopost_files.get('factures', [])
                    
                    for facture in factures_to_process:
                        facture.seek(0)
                        df_invoice = load_chronopost_invoice(facture)
                        df_invoice['Num_Facture'] = facture.name.split('_')[0] if '_' in facture.name else facture.name[:8]
                        all_invoices.append(df_invoice)
                        
                        facture.seek(0)
                        df_raw = pd.read_excel(facture, sheet_name='Table 1', header=None)
                        surplus = extract_surplus(df_raw)
                        for s in surplus:
                            s['Num_Facture'] = facture.name.split('_')[0] if '_' in facture.name else facture.name[:8]
                        all_surplus.extend(surplus)
                    
                    if not all_invoices:
                        st.error("❌ Aucune donnée extraite des factures")
                    else:
                        df_invoices = pd.concat(all_invoices, ignore_index=True)
                        
                        # Charger et fusionner fichiers logisticien (si disponibles)
                        df_log = None
                        if len(available_files) > 0:
                            all_log_data = []
                            for log_file in available_files:
                                try:
                                    log_file.seek(0)
                                    df_log_temp = pd.read_excel(log_file, sheet_name='Facturation préparation')
                                    all_log_data.append(df_log_temp)
                                except Exception as e:
                                    st.warning(f"⚠️ Erreur lecture {log_file.name}: {str(e)}")
                            
                            if all_log_data:
                                df_log = pd.concat(all_log_data, ignore_index=True)
                                df_log = df_log.drop_duplicates(subset=['Numéro de tracking'], keep='first')
                                
                                df_log = df_log.rename(columns={
                                    'Numéro de tracking': 'Tracking',
                                    'Poids expédition': 'Poids_Logisticien',
                                    'Nom du partenaire': 'Partenaire',
                                    'Pays destination': 'Pays',
                                    "Numéro de commande d'origine": 'Num_Commande_Origine',
                                    'Numéro de commande partenaire': 'Num_Commande_Partenaire'
                                })
                                df_log['Poids_Logisticien'] = df_log['Poids_Logisticien'] / 1000
                        
                        # Merge avec logisticien (si disponible)
                        if df_log is not None:
                            df = df_invoices.merge(
                                df_log[['Tracking', 'Poids_Logisticien', 'Partenaire', 'Pays', 
                                        'Num_Commande_Origine', 'Num_Commande_Partenaire']],
                                on='Tracking',
                                how='left'
                            )
                        else:
                            # Pas de fichiers logisticien - créer colonnes vides
                            df = df_invoices.copy()
                            df['Poids_Logisticien'] = None
                            df['Partenaire'] = 'Non attribué'
                            df['Pays'] = None
                            df['Num_Commande_Origine'] = None
                            df['Num_Commande_Partenaire'] = None
                            st.info("ℹ️ Analyse sans fichiers logisticiens : certaines colonnes seront vides")
                        
                        # Calculs
                        df['Prix_Theorique_HT'] = df.apply(
                            lambda row: get_theoretical_price(row['Poids_Logisticien'], row['Pays']) if pd.notna(row['Poids_Logisticien']) else None,
                            axis=1
                        )
                        
                        df['Prix_Selon_Poids_Chronopost'] = df.apply(
                            lambda row: get_theoretical_price(row['Poids_Chronopost'], row['Pays']) if pd.notna(row['Pays']) else None,
                            axis=1
                        )
                        
                        df['Difference_Prix'] = df['Prix_Facture_HT'] - df['Prix_Theorique_HT']
                        df['Ecart_Poids'] = df['Poids_Chronopost'] - df['Poids_Logisticien']
                        
                        # Surplus
                        df_surplus = pd.DataFrame(all_surplus) if all_surplus else pd.DataFrame()
                        if not df_surplus.empty and df_log is not None:
                            df_surplus = df_surplus.merge(
                                df_log[['Tracking', 'Partenaire', 'Num_Commande_Origine', 'Num_Commande_Partenaire']],
                                on='Tracking',
                                how='left'
                            )
                        elif not df_surplus.empty:
                            df_surplus['Partenaire'] = 'Non attribué'
                            df_surplus['Num_Commande_Origine'] = None
                            df_surplus['Num_Commande_Partenaire'] = None
                        
                        # Sauvegarder
                        st.session_state.chronopost_data = {
                            'df': df,
                            'df_surplus': df_surplus,
                            'timestamp': datetime.now()
                        }
                        
                        # 💾 SAUVEGARDE AUTOMATIQUE
                        persistence.save_module_files('chronopost', st.session_state.chronopost_files)
                        persistence.save_module_data('chronopost', st.session_state.chronopost_data)
                        
                        # 📚 AUTO-ARCHIVAGE DANS LA BIBLIOTHÈQUE
                        success, year, month = persistence.auto_archive_analysis(
                            'Chronopost',
                            df,
                            st.session_state.chronopost_data
                        )
                        
                        # Message avec info archivage
                        if success:
                            from modules.bibliotheque import get_month_name
                            st.success(f"✅ Analyse terminée et archivée ({get_month_name(month)} {year})")
                        else:
                            st.success("✅ Analyse terminée et sauvegardée !")
                        
                        st.rerun()
                
                except Exception as e:
                    st.error(f"❌ Erreur: {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
        
        # Bouton réinitialisation
        st.markdown("---")
        if st.button("🗑️ Réinitialiser", type="secondary", use_container_width=True):
            persistence.delete_module_data('chronopost')
            st.session_state.chronopost_files = {}
            st.session_state.chronopost_data = None
            st.session_state.chronopost_auto_loaded = False
            st.success("✅ Module réinitialisé !")
            st.rerun()
    
    # Affichage principal
    if st.session_state.chronopost_data is None:
        st.info("👈 Uploadez vos fichiers et lancez l'analyse")
        
        st.markdown("""
        ### 📋 Fichiers requis:
        1. **Factures Chronopost** (.xlsx) - Une ou plusieurs factures
        2. **Fichiers Logisticien** - Uploadés depuis la page d'accueil (partagés entre modules)
        
        ### ✨ Fonctionnalités:
        - 📊 Synthèse par partenaire
        - 💰 Analyse des surplus (7 types détectés)
        - 📋 Détail par commande
        - 🔄 Suivi des retours expéditeur
        - 💾 Sauvegarde automatique
        """)
    
    else:
        data = st.session_state.chronopost_data
        df = data['df']
        df_surplus = data['df_surplus']
        
        # Forcer la conversion des colonnes numériques après chargement
        numeric_cols_main = ['Poids_Logisticien', 'Poids_Chronopost', 'Ecart_Poids',
                             'Prix_Theorique_HT', 'Prix_Facture_HT', 'Difference_Prix']
        for col in numeric_cols_main:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        
        if df_surplus is not None and not df_surplus.empty:
            surplus_cols = ['Poids', 'Montant_Surplus']
            for col in surplus_cols:
                if col in df_surplus.columns:
                    df_surplus[col] = pd.to_numeric(df_surplus[col], errors='coerce').fillna(0)
        
        # Statistiques globales
        col1, col2, col3, col4 = st.columns(4)
        
        ecarts_defaveur = df[df['Difference_Prix'] > 0]['Difference_Prix'].sum()
        nb_ecarts_defaveur = len(df[df['Difference_Prix'] > 0])
        
        with col1:
            st.metric("Total envois", len(df))
            if 'Date' in df.columns and not df['Date'].isna().all():
                date_min = df['Date'].min()
                date_max = df['Date'].max()
                st.caption(f"📅 Du {date_min.strftime('%d/%m/%Y')} au {date_max.strftime('%d/%m/%Y')}")
        with col2:
            st.metric(
                "⚖️ Écarts poids/prix", 
                f"{ecarts_defaveur:.2f} €", 
                delta=f"{nb_ecarts_defaveur} cas en votre défaveur",
                delta_color="inverse"
            )
        with col3:
            total_surplus = df_surplus['Montant_Surplus'].sum() if not df_surplus.empty else 0
            st.metric("Total surplus", f"{total_surplus:.2f} €", delta=f"{len(df_surplus)} cas")
        with col4:
            total_contester = ecarts_defaveur + total_surplus
            st.metric("💸 TOTAL À CONTESTER", f"{total_contester:.2f} €", delta_color="inverse")
        
        st.markdown("---")
        
        # TABS
        tab1, tab2, tab3, tab4 = st.tabs([
            "📊 Synthèse par Partenaire",
            "💰 Surplus par Partenaire",
            "📋 Détail par Commande",
            "🔄 Retours Expéditeur"
        ])
        
        # TAB 1: SYNTHÈSE
        with tab1:
            st.header("📊 Synthèse par Partenaire")
            
            partenaires = ['Tous'] + sorted(df['Partenaire'].dropna().unique().tolist())
            selected_partner = st.selectbox("Filtrer par partenaire:", partenaires, key="synthese_partner")
            
            df_filtered = df if selected_partner == 'Tous' else df[df['Partenaire'] == selected_partner]
            
            synthese = df_filtered.groupby('Partenaire').agg({
                'Tracking': 'count',
                'Poids_Logisticien': 'sum',
                'Poids_Chronopost': 'sum',
                'Prix_Theorique_HT': 'sum',
                'Prix_Facture_HT': 'sum',
                'Difference_Prix': 'sum'
            }).reset_index()
            
            synthese.columns = ['Partenaire', 'Nb Envois', 'Poids Log. (kg)', 'Poids Chrono (kg)', 
                               'Prix Théorique (€)', 'Prix Facturé (€)', 'Écart (€)']
            
            # Forcer la conversion en numérique pour éviter TypeError
            numeric_cols = ['Nb Envois', 'Poids Log. (kg)', 'Poids Chrono (kg)', 
                           'Prix Théorique (€)', 'Prix Facturé (€)', 'Écart (€)']
            for col in numeric_cols:
                synthese[col] = pd.to_numeric(synthese[col], errors='coerce').fillna(0)
            
            if not df_surplus.empty:
                surplus_by_partner = df_surplus.groupby('Partenaire')['Montant_Surplus'].sum().reset_index()
                surplus_by_partner.columns = ['Partenaire', 'Total Surplus (€)']
                synthese = synthese.merge(surplus_by_partner, on='Partenaire', how='left')
                synthese['Total Surplus (€)'] = pd.to_numeric(synthese['Total Surplus (€)'], errors='coerce').fillna(0)
                synthese['Total à Contester (€)'] = synthese['Écart (€)'] + synthese['Total Surplus (€)']
            
            for col in ['Poids Log. (kg)', 'Poids Chrono (kg)']:
                synthese[col] = synthese[col].round(3)
            for col in ['Prix Théorique (€)', 'Prix Facturé (€)', 'Écart (€)']:
                synthese[col] = synthese[col].round(2)
            if 'Total Surplus (€)' in synthese.columns:
                synthese['Total Surplus (€)'] = synthese['Total Surplus (€)'].round(2)
                synthese['Total à Contester (€)'] = synthese['Total à Contester (€)'].round(2)
            
            synthese = synthese.sort_values('Écart (€)', ascending=False)
            
            st.dataframe(synthese, use_container_width=True, hide_index=True)
            
            if st.button("📥 Exporter Synthèse", key="export_synthese"):
                excel_data = create_excel_export(synthese, "Synthèse Partenaires")
                st.download_button(
                    label="Télécharger Excel",
                    data=excel_data,
                    file_name=f"synthese_chronopost_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        
        # TAB 2: SURPLUS
        with tab2:
            st.header("💰 Surplus par Partenaire")
            
            if df_surplus.empty:
                st.info("Aucun surplus détecté")
            else:
                col1, col2 = st.columns(2)
                
                with col1:
                    partenaires_surplus = ['Tous'] + sorted(df_surplus['Partenaire'].dropna().unique().tolist())
                    selected_partner_surplus = st.selectbox("Filtrer par partenaire:", partenaires_surplus, key="surplus_partner")
                
                with col2:
                    all_surplus_types = sorted(df_surplus['Type_Surplus'].unique().tolist())
                    default_surplus_types = [
                        'Etiquette non conforme',
                        'Supplément manutention',
                        'Supplément hors norme',
                        'Retour expéditeur'
                    ]
                    default_selected = [t for t in default_surplus_types if t in all_surplus_types]
                    
                    selected_surplus_types = st.multiselect(
                        "Filtrer par type de surplus:",
                        options=all_surplus_types,
                        default=default_selected,
                        key="surplus_types",
                        help="Pré-sélection: Etiquette, Manutention, Hors norme, Retour expéditeur"
                    )
                
                df_surplus_filtered = df_surplus if selected_partner_surplus == 'Tous' else df_surplus[df_surplus['Partenaire'] == selected_partner_surplus]
                
                if selected_surplus_types:
                    df_surplus_filtered = df_surplus_filtered[df_surplus_filtered['Type_Surplus'].isin(selected_surplus_types)]
                else:
                    st.warning("⚠️ Aucun type de surplus sélectionné")
                    df_surplus_filtered = pd.DataFrame()
                
                if not df_surplus_filtered.empty:
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Surplus affichés", len(df_surplus_filtered))
                    with col2:
                        st.metric("Types sélectionnés", len(selected_surplus_types))
                    with col3:
                        st.metric("Montant total", f"{df_surplus_filtered['Montant_Surplus'].sum():.2f} €")
                    
                    st.subheader("📊 Répartition par type")
                    recap = df_surplus_filtered.groupby('Type_Surplus').agg({
                        'Tracking': 'count',
                        'Montant_Surplus': 'sum'
                    }).reset_index()
                    recap.columns = ['Type de Surplus', 'Nombre', 'Montant Total (€)']
                    
                    # Forcer la conversion en numérique
                    recap['Montant Total (€)'] = pd.to_numeric(recap['Montant Total (€)'], errors='coerce').fillna(0)
                    recap['Montant Total (€)'] = recap['Montant Total (€)'].round(2)
                    recap = recap.sort_values('Montant Total (€)', ascending=False)
                    
                    st.dataframe(recap, use_container_width=True, hide_index=True)
                    
                    st.subheader("📋 Détail complet")
                    df_display = df_surplus_filtered[[
                        'Date', 'Partenaire', 'Tracking', 'Num_Commande_Origine',
                        'Type_Surplus', 'Montant_Surplus'
                    ]].copy()
                    df_display.columns = ['Date', 'Partenaire', 'Tracking', 'N° Cmd Origine',
                                         'Type Surplus', 'Montant (€)']
                    df_display = df_display.sort_values(['Type Surplus', 'Partenaire', 'Date'])
                    
                    st.dataframe(df_display, use_container_width=True, hide_index=True)
                    
                    st.info(f"📌 **Filtres actifs**: Partenaire: {selected_partner_surplus} | Types: {len(selected_surplus_types)}/{len(all_surplus_types)}")
                    
                    if st.button("📥 Exporter Surplus", key="export_surplus"):
                        partner_suffix = f"_{selected_partner_surplus}" if selected_partner_surplus != 'Tous' else ""
                        types_suffix = f"_{len(selected_surplus_types)}types" if len(selected_surplus_types) < len(all_surplus_types) else ""
                        
                        excel_data = create_excel_export(df_display, "Surplus Partenaires")
                        st.download_button(
                            label="Télécharger Excel",
                            data=excel_data,
                            file_name=f"surplus_chronopost{partner_suffix}{types_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
        
        # TAB 3: DÉTAIL
        with tab3:
            st.header("📋 Détail par Commande")
            
            col1, col2 = st.columns(2)
            with col1:
                partenaires_detail = ['Tous'] + sorted(df['Partenaire'].dropna().unique().tolist())
                selected_partner_detail = st.selectbox("Filtrer par partenaire:", partenaires_detail, key="detail_partner")
            with col2:
                show_only_ecarts = st.checkbox("Afficher uniquement les écarts", value=False)
            
            df_detail = df.copy()
            if selected_partner_detail != 'Tous':
                df_detail = df_detail[df_detail['Partenaire'] == selected_partner_detail]
            if show_only_ecarts:
                df_detail = df_detail[
                    (df_detail['Difference_Prix'] > 0) | 
                    (df_detail['Ecart_Poids'].abs() > 0.01)
                ]
            
            df_detail_display = df_detail[[
                'Date', 'Partenaire', 'Tracking', 'Num_Commande_Origine', 'Pays',
                'Poids_Logisticien', 'Poids_Chronopost', 'Ecart_Poids',
                'Prix_Theorique_HT', 'Prix_Facture_HT', 'Difference_Prix'
            ]].copy()
            
            df_detail_display.columns = [
                'Date', 'Partenaire', 'Tracking', 'N° Commande', 'Pays',
                'Poids Log. (kg)', 'Poids Chrono (kg)', 'Écart Poids (kg)',
                'Prix Théo. (€)', 'Prix Facturé (€)', 'Écart Prix (€)'
            ]
            
            # Forcer la conversion en numérique pour éviter TypeError
            numeric_cols_detail = ['Poids Log. (kg)', 'Poids Chrono (kg)', 'Écart Poids (kg)',
                                   'Prix Théo. (€)', 'Prix Facturé (€)', 'Écart Prix (€)']
            for col in numeric_cols_detail:
                df_detail_display[col] = pd.to_numeric(df_detail_display[col], errors='coerce').fillna(0)
            
            for col in ['Poids Log. (kg)', 'Poids Chrono (kg)', 'Écart Poids (kg)']:
                df_detail_display[col] = df_detail_display[col].round(3)
            for col in ['Prix Théo. (€)', 'Prix Facturé (€)', 'Écart Prix (€)']:
                df_detail_display[col] = df_detail_display[col].round(2)
            
            df_detail_display = df_detail_display.sort_values('Écart Prix (€)', ascending=False)
            
            st.dataframe(df_detail_display, use_container_width=True, hide_index=True)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Envois affichés", len(df_detail_display))
            with col2:
                ecarts_pos = len(df_detail_display[df_detail_display['Écart Prix (€)'] > 0])
                st.metric("Avec écart prix", ecarts_pos)
            with col3:
                ecarts_poids = len(df_detail_display[df_detail_display['Écart Poids (kg)'].abs() > 0.01])
                st.metric("Avec écart poids", ecarts_poids)
            
            if st.button("📥 Exporter Détail", key="export_detail"):
                df_export_filtered = df_detail_display[df_detail_display['Écart Prix (€)'] > 0].copy()
                
                if len(df_export_filtered) == 0:
                    st.warning("⚠️ Aucun écart en votre défaveur à exporter")
                else:
                    excel_data = create_excel_export(df_export_filtered, "Détail Commandes")
                    st.download_button(
                        label=f"Télécharger Excel ({len(df_export_filtered)} écarts en défaveur)",
                        data=excel_data,
                        file_name=f"detail_chronopost_ecarts_defaveur_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
        
        # TAB 4: RETOURS
        with tab4:
            st.header("🔄 Retours Expéditeur")
            
            if df_surplus.empty:
                st.info("Aucun retour détecté")
            else:
                df_retours = df_surplus[
                    df_surplus['Type_Surplus'].isin(['Retour expéditeur', 'Traitement Retour expéditeur'])
                ].copy()
                
                if df_retours.empty:
                    st.info("Aucun retour expéditeur détecté")
                else:
                    partenaires_retours = ['Tous'] + sorted(df_retours['Partenaire'].dropna().unique().tolist())
                    selected_partner_retours = st.selectbox("Filtrer par partenaire:", partenaires_retours, key="retours_partner")
                    
                    df_retours_filtered = df_retours if selected_partner_retours == 'Tous' else df_retours[df_retours['Partenaire'] == selected_partner_retours]
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total retours", len(df_retours_filtered))
                    with col2:
                        retours_complets = len(df_retours_filtered[df_retours_filtered['Type_Surplus'] == 'Retour expéditeur'])
                        st.metric("Retours complets", retours_complets, delta=f"{retours_complets * 20:.0f} €")
                    with col3:
                        traitements = len(df_retours_filtered[df_retours_filtered['Type_Surplus'] == 'Traitement Retour expéditeur'])
                        st.metric("Traitements", traitements, delta=f"{traitements * 2:.0f} €")
                    
                    st.subheader("📊 Retours par partenaire")
                    retours_by_partner = df_retours_filtered.groupby('Partenaire').agg({
                        'Tracking': 'count',
                        'Montant_Surplus': 'sum'
                    }).reset_index()
                    retours_by_partner.columns = ['Partenaire', 'Nombre de Retours', 'Coût Total (€)']
                    retours_by_partner = retours_by_partner.sort_values('Coût Total (€)', ascending=False)
                    
                    st.dataframe(retours_by_partner, use_container_width=True, hide_index=True)
                    
                    st.subheader("📋 Détail des retours")
                    df_retours_display = df_retours_filtered[[
                        'Date', 'Partenaire', 'Tracking', 'Num_Commande_Origine',
                        'Type_Surplus', 'Montant_Surplus'
                    ]].copy()
                    df_retours_display.columns = ['Date', 'Partenaire', 'Tracking', 'N° Commande',
                                                  'Type de Retour', 'Coût (€)']
                    
                    st.dataframe(df_retours_display, use_container_width=True, hide_index=True)
                    
                    if st.button("📥 Exporter Retours", key="export_retours"):
                        excel_data = create_excel_export(df_retours_display, "Retours Expéditeur")
                        st.download_button(
                            label="Télécharger Excel",
                            data=excel_data,
                            file_name=f"retours_chronopost_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )

if __name__ == "__main__":
    run()
