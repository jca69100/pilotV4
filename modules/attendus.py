"""
Module : Gestion des Attendus de Réception
Permet l'import CSV et saisie manuelle du conditionnement d'entrepôt
"""

import streamlit as st
import pandas as pd
from datetime import datetime
from shared import persistence
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Liste des clients prédéfinis
CLIENTS_PREDEFINIS = [
    "AGST DEVELOPPEMENT",
    "ASTERIA INTERNATIONAL",
    "BIVOUAK",
    "BLUE SKY WELL",
    "CA VA BARBER",
    "CDM",
    "COCCINELLE SAS",
    "EDITIONS OPTIMIST",
    "GP WORKERS",
    "INTELIMMED CONSULTING",
    "INTOM ART",
    "JEJUSKIN",
    "MADEMOISELLE SAINT GERMAIN",
    "OPS ECO INNOVATION",
    "RAMDAM SOCIAL",
    "UN BATTEMENT D'AILE",
    "WRSTL FRANCE"
]

def load_csv_data():
    """Charge les données CSV depuis la persistence"""
    data = persistence.load_module_data('attendus_csv')
    if data and 'df' in data:
        return data['df']
    return pd.DataFrame(columns=['client', 'attendu', 'sku', 'quantite'])

def save_csv_data(df):
    """Sauvegarde les données CSV"""
    return persistence.save_module_data('attendus_csv', {'df': df})

def load_manual_data():
    """Charge les données manuelles depuis la persistence"""
    data = persistence.load_module_data('attendus_manual')
    if data and 'df' in data:
        return data['df']
    return pd.DataFrame(columns=['id', 'client', 'attendu', 'palette_multi', 'palette_mono', 'carton', 'timestamp'])

def save_manual_data(df):
    """Sauvegarde les données manuelles"""
    return persistence.save_module_data('attendus_manual', {'df': df})

def load_custom_clients():
    """Charge la liste des clients personnalisés"""
    data = persistence.load_module_data('attendus_custom_clients')
    if data and 'clients' in data:
        return data['clients']
    return []

def save_custom_clients(clients_list):
    """Sauvegarde la liste des clients personnalisés"""
    return persistence.save_module_data('attendus_custom_clients', {'clients': clients_list})

def add_custom_client(client_name):
    """Ajoute un nouveau client personnalisé"""
    custom_clients = load_custom_clients()
    if client_name not in custom_clients and client_name not in CLIENTS_PREDEFINIS:
        custom_clients.append(client_name)
        custom_clients.sort()  # Trier par ordre alphabétique
        save_custom_clients(custom_clients)
        return True
    return False

def get_all_clients():
    """Retourne la liste complète des clients (prédéfinis + personnalisés)"""
    custom_clients = load_custom_clients()
    all_clients = CLIENTS_PREDEFINIS + custom_clients
    return sorted(all_clients)

def parse_csv_file(uploaded_file):
    """
    Parse le fichier CSV et extrait les colonnes requises
    
    Colonnes attendues:
    - resellers.name → client
    - source_ref → attendu
    - supply_capsule_items.source_ref → sku
    - supply_capsule_items.received_quantity → quantite
    """
    try:
        df = pd.read_csv(uploaded_file)
        
        # Colonnes requises
        required_cols = {
            'resellers.name': 'client',
            'source_ref': 'attendu',
            'supply_capsule_items.source_ref': 'sku',
            'supply_capsule_items.received_quantity': 'quantite'
        }
        
        # Vérifier que toutes les colonnes sont présentes
        missing_cols = [col for col in required_cols.keys() if col not in df.columns]
        if missing_cols:
            return None, f"Colonnes manquantes : {', '.join(missing_cols)}"
        
        # Extraire et renommer les colonnes
        df_parsed = df[list(required_cols.keys())].copy()
        df_parsed.columns = list(required_cols.values())
        
        # Nettoyer les données
        df_parsed['quantite'] = pd.to_numeric(df_parsed['quantite'], errors='coerce').fillna(0).astype(int)
        df_parsed = df_parsed[df_parsed['quantite'] > 0]  # Supprimer les lignes avec quantité 0
        
        return df_parsed, None
        
    except Exception as e:
        return None, f"Erreur lors de la lecture du CSV : {str(e)}"

def aggregate_data(df_csv, df_manual):
    """Fusionne et agrège les données CSV et manuelles"""
    
    aggregated = []
    
    # Agréger CSV par (client, attendu)
    if len(df_csv) > 0:
        csv_grouped = df_csv.groupby(['client', 'attendu']).agg({
            'sku': 'nunique'
        }).reset_index()
        csv_grouped.columns = ['client', 'attendu', 'nb_sku']
        
        for _, row in csv_grouped.iterrows():
            aggregated.append({
                'client': row['client'],
                'attendu': row['attendu'],
                'palette_multi': 0,
                'palette_mono': 0,
                'carton': 0,
                'nb_sku': row['nb_sku'],
                'source': 'CSV'
            })
    
    # Agréger Manuel par (client, attendu)
    if len(df_manual) > 0:
        manual_grouped = df_manual.groupby(['client', 'attendu']).agg({
            'palette_multi': 'sum',
            'palette_mono': 'sum',
            'carton': 'sum'
        }).reset_index()
        
        for _, row in manual_grouped.iterrows():
            # Chercher si déjà dans aggregated (même client/attendu du CSV)
            found = False
            for item in aggregated:
                if item['client'] == row['client'] and item['attendu'] == row['attendu']:
                    item['palette_multi'] = row['palette_multi']
                    item['palette_mono'] = row['palette_mono']
                    item['carton'] = row['carton']
                    item['source'] = 'CSV + Manuel'
                    found = True
                    break
            
            if not found:
                aggregated.append({
                    'client': row['client'],
                    'attendu': row['attendu'],
                    'palette_multi': row['palette_multi'],
                    'palette_mono': row['palette_mono'],
                    'carton': row['carton'],
                    'nb_sku': 0,
                    'source': 'Manuel'
                })
    
    df_synthese = pd.DataFrame(aggregated)
    
    if len(df_synthese) == 0:
        df_synthese = pd.DataFrame(columns=['client', 'attendu', 'palette_multi', 'palette_mono', 'carton', 'nb_sku', 'source'])
    
    return df_synthese

def calculate_kpi(df_synthese):
    """Calcule les KPI globaux"""
    if len(df_synthese) == 0:
        return {
            'total_palettes': 0,
            'total_cartons': 0,
            'nb_clients': 0
        }
    
    return {
        'total_palettes': int(df_synthese['palette_multi'].sum() + df_synthese['palette_mono'].sum()),
        'total_cartons': int(df_synthese['carton'].sum()),
        'nb_clients': df_synthese['client'].nunique()
    }

def export_synthese_excel(df, client_filter=None):
    """Exporte la synthèse en Excel"""
    
    # Appliquer le filtre si nécessaire
    if client_filter and client_filter != "Tous":
        df = df[df['client'] == client_filter].copy()
    
    # Créer un fichier Excel en mémoire
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Synthèse', index=False)
        
        # Formater le fichier
        workbook = writer.book
        worksheet = writer.sheets['Synthèse']
        
        # En-têtes en gras
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajuster les largeurs de colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def export_detail_excel(df_csv, client_filter=None, attendu_filter=None):
    """Exporte le détail en Excel"""
    
    df = df_csv.copy()
    
    # Appliquer les filtres
    if client_filter and client_filter != "Tous":
        df = df[df['client'] == client_filter]
    
    if attendu_filter and attendu_filter != "Tous":
        df = df[df['attendu'] == attendu_filter]
    
    # Créer un fichier Excel en mémoire
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Détail', index=False)
        
        # Formater le fichier
        workbook = writer.book
        worksheet = writer.sheets['Détail']
        
        # En-têtes en gras
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajuster les largeurs de colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def display_kpi(df_synthese):
    """Affiche les 3 KPI en haut de page"""
    kpi = calculate_kpi(df_synthese)
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric(
            label="📦 Total Palettes",
            value=f"{kpi['total_palettes']:,}".replace(',', ' ')
        )
    
    with col2:
        st.metric(
            label="📋 Total Cartons",
            value=f"{kpi['total_cartons']:,}".replace(',', ' ')
        )
    
    with col3:
        st.metric(
            label="👥 Nombre de Clients",
            value=kpi['nb_clients']
        )

def display_csv_import():
    """Affiche la section d'import CSV"""
    st.markdown("---")
    st.markdown("### 📁 Import CSV")
    
    uploaded_file = st.file_uploader(
        "Choisir un fichier CSV",
        type=['csv'],
        help="Format attendu : colonnes resellers.name, source_ref, supply_capsule_items.source_ref, supply_capsule_items.received_quantity"
    )
    
    if uploaded_file is not None:
        df_parsed, error = parse_csv_file(uploaded_file)
        
        if error:
            st.error(f"❌ {error}")
        else:
            st.success(f"✅ Fichier importé avec succès ! **{len(df_parsed)} lignes** traitées")
            
            # Aperçu
            with st.expander("📋 Aperçu des données", expanded=False):
                st.dataframe(df_parsed.head(10), use_container_width=True)
                st.caption(f"Affichage des 10 premières lignes sur {len(df_parsed)}")
            
            # Bouton pour sauvegarder
            if st.button("💾 Enregistrer les données CSV", type="primary"):
                save_csv_data(df_parsed)
                st.success("✅ Données CSV sauvegardées !")
                st.rerun()

def display_manual_form(df_manual):
    """Affiche le formulaire de saisie manuelle"""
    st.markdown("---")
    st.markdown("### ✍️ Saisie Manuelle de Conditionnement")
    
    # Vérifier si on est en mode édition
    edit_mode = st.session_state.get('edit_mode_id') is not None
    
    if edit_mode:
        st.warning(f"⚠️ **MODE MODIFICATION** - ID: {st.session_state.edit_mode_id}")
    
    # Obtenir la liste complète des clients
    all_clients = get_all_clients()
    clients_options = all_clients + ["➕ Nouveau client"]
    
    # Formulaire
    col1, col2 = st.columns(2)
    
    with col1:
        # Client
        client_index = 0
        selected_client = None
        
        if edit_mode:
            # Pré-remplir avec les données existantes
            edit_data = df_manual[df_manual['id'] == st.session_state.edit_mode_id].iloc[0]
            if edit_data['client'] in all_clients:
                client_index = all_clients.index(edit_data['client'])
                selected_client = edit_data['client']
        elif 'last_selected_client' in st.session_state:
            if st.session_state.last_selected_client in all_clients:
                client_index = all_clients.index(st.session_state.last_selected_client)
                selected_client = st.session_state.last_selected_client
        
        client_choice = st.selectbox(
            "Client",
            options=clients_options,
            index=client_index,
            key="attendu_client_choice"
        )
        
        # Si "Nouveau client" sélectionné, afficher champ de saisie
        if client_choice == "➕ Nouveau client":
            new_client_name = st.text_input(
                "Nom du nouveau client",
                key="new_client_name",
                help="Saisir le nom du nouveau client (sera ajouté à la liste)"
            )
            client = new_client_name.strip().upper() if new_client_name else ""
        else:
            client = client_choice
    
    with col2:
        # Attendu
        attendu_value = ""
        if edit_mode:
            attendu_value = edit_data['attendu']
        
        attendu = st.text_input(
            "Nom de l'attendu",
            value=attendu_value,
            key="attendu_name"
        )
    
    # Message d'aide pour nouveau client
    if client_choice == "➕ Nouveau client":
        st.info("💡 Le nouveau client sera automatiquement ajouté à la liste après enregistrement")
    
    # Quantités
    col1, col2, col3 = st.columns(3)
    
    with col1:
        palette_multi_value = 0
        if edit_mode:
            palette_multi_value = int(edit_data['palette_multi'])
        
        palette_multi = st.number_input(
            "Palettes Multi-Réf",
            min_value=0,
            value=palette_multi_value,
            step=1,
            key="palette_multi"
        )
    
    with col2:
        palette_mono_value = 0
        if edit_mode:
            palette_mono_value = int(edit_data['palette_mono'])
        
        palette_mono = st.number_input(
            "Palettes Mono-Réf",
            min_value=0,
            value=palette_mono_value,
            step=1,
            key="palette_mono"
        )
    
    with col3:
        carton_value = 0
        if edit_mode:
            carton_value = int(edit_data['carton'])
        
        carton = st.number_input(
            "Cartons",
            min_value=0,
            value=carton_value,
            step=1,
            key="carton"
        )
    
    # Boutons
    col1, col2, col3 = st.columns([2, 2, 6])
    
    with col1:
        if st.button("💾 Enregistrer", type="primary", use_container_width=True):
            # Validation
            if not client:
                st.error("⚠️ Veuillez sélectionner ou saisir un nom de client")
            elif not attendu:
                st.error("⚠️ Veuillez saisir un nom d'attendu")
            elif palette_multi == 0 and palette_mono == 0 and carton == 0:
                st.error("⚠️ Au moins une quantité doit être supérieure à 0")
            else:
                # Si nouveau client, l'ajouter à la liste
                if client_choice == "➕ Nouveau client" and client:
                    if add_custom_client(client):
                        st.success(f"✅ Nouveau client **{client}** ajouté à la liste !")
                
                # Sauvegarder le client pour la prochaine fois
                st.session_state.last_selected_client = client
                
                if edit_mode:
                    # Mode modification - mettre à jour l'entrée existante
                    df_manual.loc[df_manual['id'] == st.session_state.edit_mode_id, 'client'] = client
                    df_manual.loc[df_manual['id'] == st.session_state.edit_mode_id, 'attendu'] = attendu
                    df_manual.loc[df_manual['id'] == st.session_state.edit_mode_id, 'palette_multi'] = palette_multi
                    df_manual.loc[df_manual['id'] == st.session_state.edit_mode_id, 'palette_mono'] = palette_mono
                    df_manual.loc[df_manual['id'] == st.session_state.edit_mode_id, 'carton'] = carton
                    df_manual.loc[df_manual['id'] == st.session_state.edit_mode_id, 'timestamp'] = datetime.now()
                    
                    save_manual_data(df_manual)
                    st.success("✅ Entrée modifiée avec succès !")
                    
                    # Sortir du mode édition
                    st.session_state.edit_mode_id = None
                    st.rerun()
                else:
                    # Mode création - nouvelle entrée
                    new_entry = pd.DataFrame([{
                        'id': int(datetime.now().timestamp() * 1000),
                        'client': client,
                        'attendu': attendu,
                        'palette_multi': palette_multi,
                        'palette_mono': palette_mono,
                        'carton': carton,
                        'timestamp': datetime.now()
                    }])
                    
                    df_manual = pd.concat([df_manual, new_entry], ignore_index=True)
                    save_manual_data(df_manual)
                    st.success(f"✅ Entrée ajoutée pour **{client}** - **{attendu}**")
                    st.rerun()
    
    with col2:
        if edit_mode:
            if st.button("❌ Annuler", use_container_width=True):
                st.session_state.edit_mode_id = None
                st.rerun()

def display_synthese(df_synthese, df_manual):
    """Affiche le tableau de synthèse"""
    st.markdown("### 📊 Synthèse par Client")
    
    # Filtre client
    clients_list = ["Tous"] + sorted(df_synthese['client'].unique().tolist()) if len(df_synthese) > 0 else ["Tous"]
    client_filter = st.selectbox("Filtrer par client", clients_list, key="synthese_client_filter")
    
    # Appliquer le filtre
    if client_filter != "Tous":
        df_filtered = df_synthese[df_synthese['client'] == client_filter].copy()
    else:
        df_filtered = df_synthese.copy()
    
    if len(df_filtered) == 0:
        st.info("ℹ️ Aucune donnée à afficher")
        return
    
    # Préparer l'affichage
    df_display = df_filtered[['client', 'attendu', 'palette_multi', 'palette_mono', 'carton', 'nb_sku', 'source']].copy()
    df_display.columns = ['Client', 'Attendu', 'Palettes Multi', 'Palettes Mono', 'Cartons', 'Nb SKU', 'Source']
    
    # Ajouter colonne Actions pour les entrées manuelles
    def get_actions(row):
        # Chercher si cette entrée a une source manuelle
        client = row['Client']
        attendu = row['Attendu']
        source = row['Source']
        
        if 'Manuel' in source:
            # Trouver l'ID de l'entrée manuelle
            manual_entries = df_manual[(df_manual['client'] == client) & (df_manual['attendu'] == attendu)]
            if len(manual_entries) > 0:
                entry_id = manual_entries.iloc[0]['id']
                return f"edit_{entry_id}"
        return ""
    
    df_display['Actions'] = df_display.apply(get_actions, axis=1)
    
    # Afficher le tableau
    st.dataframe(
        df_display,
        use_container_width=True,
        hide_index=True,
        column_config={
            'Client': st.column_config.TextColumn('Client', width='medium'),
            'Attendu': st.column_config.TextColumn('Attendu', width='medium'),
            'Palettes Multi': st.column_config.NumberColumn('Palettes Multi', width='small'),
            'Palettes Mono': st.column_config.NumberColumn('Palettes Mono', width='small'),
            'Cartons': st.column_config.NumberColumn('Cartons', width='small'),
            'Nb SKU': st.column_config.NumberColumn('Nb SKU', width='small'),
            'Source': st.column_config.TextColumn('Source', width='small'),
            'Actions': st.column_config.TextColumn('Actions', width='small')
        }
    )
    
    # Afficher les boutons d'action
    for _, row in df_display.iterrows():
        if row['Actions']:
            entry_id = int(row['Actions'].replace('edit_', ''))
            col1, col2, col3 = st.columns([6, 1, 1])
            
            with col1:
                st.write(f"**{row['Client']}** - {row['Attendu']}")
            
            with col2:
                if st.button(f"✏️", key=f"edit_{entry_id}"):
                    st.session_state.edit_mode_id = entry_id
                    st.rerun()
            
            with col3:
                if st.button(f"🗑️", key=f"delete_{entry_id}"):
                    if st.session_state.get(f'confirm_delete_{entry_id}'):
                        # Supprimer
                        df_manual_updated = df_manual[df_manual['id'] != entry_id]
                        save_manual_data(df_manual_updated)
                        st.success("✅ Entrée supprimée")
                        del st.session_state[f'confirm_delete_{entry_id}']
                        st.rerun()
                    else:
                        st.session_state[f'confirm_delete_{entry_id}'] = True
                        st.warning("⚠️ Cliquer à nouveau pour confirmer")
    
    # Ligne de total
    st.markdown("---")
    total_label = f"TOTAL ({client_filter})" if client_filter != "Tous" else "TOTAL (GÉNÉRAL)"
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric(total_label, "")
    
    with col2:
        st.metric("Palettes Multi", int(df_filtered['palette_multi'].sum()))
    
    with col3:
        st.metric("Palettes Mono", int(df_filtered['palette_mono'].sum()))
    
    with col4:
        st.metric("Cartons", int(df_filtered['carton'].sum()))
    
    with col5:
        st.metric("SKU", int(df_filtered['nb_sku'].sum()))
    
    # Export
    st.markdown("---")
    if st.button("📥 Exporter Synthèse en Excel", use_container_width=True):
        excel_data = export_synthese_excel(df_synthese, client_filter)
        st.download_button(
            label="💾 Télécharger",
            data=excel_data,
            file_name=f"attendus_synthese_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

def display_detail(df_csv):
    """Affiche le tableau de détail"""
    st.markdown("### 🔍 Détail par SKU")
    
    if len(df_csv) == 0:
        st.info("ℹ️ Aucune donnée CSV importée")
        return
    
    # Filtres
    col1, col2 = st.columns(2)
    
    with col1:
        clients_list = ["Tous"] + sorted(df_csv['client'].unique().tolist())
        client_filter = st.selectbox("Filtrer par client", clients_list, key="detail_client_filter")
    
    with col2:
        if client_filter != "Tous":
            attendus_list = ["Tous"] + sorted(df_csv[df_csv['client'] == client_filter]['attendu'].unique().tolist())
        else:
            attendus_list = ["Tous"] + sorted(df_csv['attendu'].unique().tolist())
        
        attendu_filter = st.selectbox("Filtrer par attendu", attendus_list, key="detail_attendu_filter")
    
    # Appliquer les filtres
    df_filtered = df_csv.copy()
    
    if client_filter != "Tous":
        df_filtered = df_filtered[df_filtered['client'] == client_filter]
    
    if attendu_filter != "Tous":
        df_filtered = df_filtered[df_filtered['attendu'] == attendu_filter]
    
    # Afficher
    st.dataframe(
        df_filtered,
        use_container_width=True,
        hide_index=True,
        column_config={
            'client': st.column_config.TextColumn('Client', width='medium'),
            'attendu': st.column_config.TextColumn('Attendu', width='medium'),
            'sku': st.column_config.TextColumn('SKU', width='medium'),
            'quantite': st.column_config.NumberColumn('Quantité', width='small')
        }
    )
    
    st.caption(f"**{len(df_filtered)}** ligne(s) affichée(s)")
    
    # Export
    st.markdown("---")
    if st.button("📥 Exporter Détail en Excel", use_container_width=True):
        excel_data = export_detail_excel(df_csv, client_filter, attendu_filter)
        st.download_button(
            label="💾 Télécharger",
            data=excel_data,
            file_name=f"attendus_detail_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

def run():
    """Point d'entrée du module"""
    st.title("📦 Gestion des Attendus de Réception")
    
    st.markdown("""
    ### 📋 À propos
    
    Ce module permet de gérer les attendus de réception d'entrepôt en combinant :
    - 📄 **Import CSV** : Données issues du système d'information
    - ✍️ **Saisie manuelle** : Informations de conditionnement (palettes, cartons)
    - 📊 **Synthèse** : Vue consolidée par client et attendu
    - 🔍 **Détail** : Vue ligne par ligne des SKU
    """)
    
    # Initialisation session state
    if 'edit_mode_id' not in st.session_state:
        st.session_state.edit_mode_id = None
    if 'last_selected_client' not in st.session_state:
        st.session_state.last_selected_client = get_all_clients()[0] if get_all_clients() else CLIENTS_PREDEFINIS[0]
    
    # Charger les données
    df_csv = load_csv_data()
    df_manual = load_manual_data()
    
    # Agréger
    df_synthese = aggregate_data(df_csv, df_manual)
    
    # KPI
    display_kpi(df_synthese)
    
    # Section de gestion des clients personnalisés
    with st.expander("👥 Gestion des Clients Personnalisés"):
        st.markdown("#### Liste des Clients")
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            st.markdown("**17 clients prédéfinis** (toujours disponibles)")
            st.caption(", ".join(CLIENTS_PREDEFINIS))
        
        custom_clients = load_custom_clients()
        
        if custom_clients:
            st.markdown("---")
            st.markdown(f"**{len(custom_clients)} client(s) personnalisé(s)**")
            
            for client in custom_clients:
                col1, col2 = st.columns([4, 1])
                with col1:
                    st.write(f"• {client}")
                with col2:
                    if st.button("🗑️", key=f"delete_client_{client}"):
                        # Vérifier si ce client est utilisé
                        is_used_manual = len(df_manual[df_manual['client'] == client]) > 0
                        is_used_csv = len(df_csv[df_csv['client'] == client]) > 0
                        
                        if is_used_manual or is_used_csv:
                            st.error(f"❌ Impossible de supprimer **{client}** : utilisé dans des attendus existants")
                        else:
                            custom_clients.remove(client)
                            save_custom_clients(custom_clients)
                            st.success(f"✅ Client **{client}** supprimé")
                            st.rerun()
        else:
            st.info("ℹ️ Aucun client personnalisé pour le moment. Utilisez '➕ Nouveau client' dans le formulaire pour en ajouter.")
        
        st.markdown("---")
        st.markdown("#### Ajouter un Client Rapidement")
        
        col1, col2 = st.columns([3, 1])
        
        with col1:
            quick_client_name = st.text_input(
                "Nom du client",
                key="quick_add_client",
                help="Le nom sera automatiquement converti en majuscules"
            )
        
        with col2:
            st.write("")  # Alignement
            st.write("")  # Alignement
            if st.button("➕ Ajouter", type="primary", use_container_width=True):
                if quick_client_name:
                    client_upper = quick_client_name.strip().upper()
                    if add_custom_client(client_upper):
                        st.success(f"✅ Client **{client_upper}** ajouté !")
                        st.rerun()
                    else:
                        st.warning("⚠️ Ce client existe déjà dans la liste")
                else:
                    st.error("⚠️ Veuillez saisir un nom de client")
    
    # Import CSV
    display_csv_import()
    
    # Formulaire saisie
    display_manual_form(df_manual)
    
    st.markdown("---")
    
    # Tabs
    tab1, tab2 = st.tabs(["📊 Synthèse par Client", "🔍 Détail par SKU"])
    
    with tab1:
        display_synthese(df_synthese, df_manual)
    
    with tab2:
        display_detail(df_csv)
    
    # Réinitialisation (Danger Zone)
    st.markdown("---")
    with st.expander("🗑️ Danger Zone - Réinitialisation"):
        st.warning("""
        ⚠️ **Attention** : Cette action supprimera **TOUTES** les données (CSV et manuelles).
        Cette action est **irréversible**.
        """)
        
        col1, col2 = st.columns([3, 1])
        
        with col2:
            if st.button("🗑️ Réinitialiser tout", type="secondary", use_container_width=True):
                if st.session_state.get('confirm_reset'):
                    # Réinitialiser
                    save_csv_data(pd.DataFrame(columns=['client', 'attendu', 'sku', 'quantite']))
                    save_manual_data(pd.DataFrame(columns=['id', 'client', 'attendu', 'palette_multi', 'palette_mono', 'carton', 'timestamp']))
                    st.success("✅ Toutes les données ont été réinitialisées")
                    del st.session_state.confirm_reset
                    st.rerun()
                else:
                    st.session_state.confirm_reset = True
                    st.error("⚠️ Cliquer à nouveau pour confirmer la réinitialisation")
