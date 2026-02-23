"""
Module : Suivi des Indemnisations Transporteurs
Permet de suivre les indemnisations reçues par transporteur et par partenaire
Version 1.0
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from shared import persistence

def get_info_from_tracking(tracking):
    """
    Recherche les informations d'une commande à partir du tracking
    dans tous les fichiers logisticien partagés
    
    Returns:
        dict: {'partenaire': str, 'num_commande_origine': str, 'num_commande_partenaire': str, 'found': bool}
    """
    tracking_clean = str(tracking).strip()
    
    # Charger fichiers logisticiens depuis la bibliothèque
    from modules.logisticiens_library import load_logisticien_files_for_analysis
    
    log_files = load_logisticien_files_for_analysis(nb_months=6)
    
    if len(log_files) == 0:
        return {
            'partenaire': 'Non trouvé',
            'num_commande_origine': '',
            'num_commande_partenaire': '',
            'found': False
        }
    
    # Parcourir tous les fichiers logisticien
    for log_file in log_files:
        try:
            # Lire le fichier Excel
            log_file.seek(0)
            
            # Essayer de lire la feuille "Facturation préparation"
            try:
                df = pd.read_excel(log_file, sheet_name="Facturation préparation")
            except:
                df = pd.read_excel(log_file)
            
            # Chercher le tracking dans la colonne "Numéro de tracking"
            if 'Numéro de tracking' in df.columns:
                # Nettoyer les trackings
                df['Tracking_Clean'] = df['Numéro de tracking'].astype(str).str.strip()
                
                # Chercher la correspondance
                match = df[df['Tracking_Clean'] == tracking_clean]
                
                if len(match) > 0:
                    row = match.iloc[0]
                    return {
                        'partenaire': row.get('Nom du partenaire', 'Non trouvé'),
                        'num_commande_origine': str(row.get("Numéro de commande d'origine", '')),
                        'num_commande_partenaire': str(row.get('Numéro de commande partenaire', '')),
                        'found': True
                    }
        except Exception as e:
            continue
    
    # Si pas trouvé
    return {
        'partenaire': 'Non trouvé',
        'num_commande_origine': '',
        'num_commande_partenaire': '',
        'found': False
    }

def export_indemnisations_excel(df):
    """Exporter les indemnisations en Excel avec mise en forme"""
    
    output = BytesIO()
    df.to_excel(output, index=False, sheet_name='Indemnisations', engine='openpyxl')
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb['Indemnisations']
    
    # Couleurs GREENLOG
    bleu_greenlog = PatternFill(start_color='2D3E50', end_color='2D3E50', fill_type='solid')
    vert_clair = PatternFill(start_color='E8F5F1', end_color='E8F5F1', fill_type='solid')
    
    # Mise en forme des en-têtes
    for cell in ws[1]:
        cell.fill = bleu_greenlog
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Alterner les couleurs de fond
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = vert_clair
    
    # Ajuster largeur des colonnes
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Créer onglet de synthèse
    ws_synthese = wb.create_sheet('Synthèse', 0)
    
    # Synthèse par partenaire et transporteur
    synthese_partner_transp = df.groupby(['Partenaire', 'Transporteur'])['Montant'].sum().reset_index()
    synthese_partner_transp = synthese_partner_transp.sort_values(['Partenaire', 'Montant'], ascending=[True, False])
    
    # Synthèse par partenaire
    synthese_partner = df.groupby('Partenaire').agg({
        'Montant': 'sum',
        'Date': 'count'
    }).reset_index()
    synthese_partner.columns = ['Partenaire', 'Total_Indemnisations', 'Nb_Cas']
    synthese_partner = synthese_partner.sort_values('Total_Indemnisations', ascending=False)
    
    # Synthèse par transporteur
    synthese_transp = df.groupby('Transporteur').agg({
        'Montant': 'sum',
        'Date': 'count'
    }).reset_index()
    synthese_transp.columns = ['Transporteur', 'Total_Indemnisations', 'Nb_Cas']
    synthese_transp = synthese_transp.sort_values('Total_Indemnisations', ascending=False)
    
    # Titre
    ws_synthese['A1'] = "SYNTHÈSE DES INDEMNISATIONS"
    ws_synthese['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws_synthese['A1'].fill = bleu_greenlog
    ws_synthese.merge_cells('A1:D1')
    ws_synthese.row_dimensions[1].height = 30
    
    ws_synthese['A2'] = f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_synthese['A3'] = f"Période : {df['Date'].min()} au {df['Date'].max()}"
    
    # Total général
    total_general = df['Montant'].sum()
    nb_total = len(df)
    ws_synthese['A5'] = "TOTAL GÉNÉRAL"
    ws_synthese['A5'].font = Font(bold=True, size=12)
    ws_synthese['B5'] = f"{total_general:.2f} €"
    ws_synthese['B5'].font = Font(bold=True, size=12, color='CC0000')
    ws_synthese['C5'] = f"{nb_total} indemnisation(s)"
    
    # Par partenaire
    row = 8
    ws_synthese[f'A{row}'] = "PAR PARTENAIRE"
    ws_synthese[f'A{row}'].font = Font(bold=True, size=11)
    ws_synthese[f'A{row}'].fill = bleu_greenlog
    ws_synthese[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    
    row += 1
    for col_idx, col_name in enumerate(['Partenaire', 'Total', 'Nb Cas'], 1):
        cell = ws_synthese.cell(row=row, column=col_idx, value=col_name)
        cell.fill = bleu_greenlog
        cell.font = Font(bold=True, color='FFFFFF')
    
    row += 1
    for _, partner_row in synthese_partner.iterrows():
        ws_synthese[f'A{row}'] = partner_row['Partenaire']
        ws_synthese[f'B{row}'] = f"{partner_row['Total_Indemnisations']:.2f} €"
        ws_synthese[f'C{row}'] = int(partner_row['Nb_Cas'])
        if row % 2 == 0:
            for col_idx in range(1, 4):
                ws_synthese.cell(row=row, column=col_idx).fill = vert_clair
        row += 1
    
    # Par transporteur
    row += 2
    ws_synthese[f'A{row}'] = "PAR TRANSPORTEUR"
    ws_synthese[f'A{row}'].font = Font(bold=True, size=11)
    ws_synthese[f'A{row}'].fill = bleu_greenlog
    ws_synthese[f'A{row}'].font = Font(bold=True, color='FFFFFF')
    
    row += 1
    for col_idx, col_name in enumerate(['Transporteur', 'Total', 'Nb Cas'], 1):
        cell = ws_synthese.cell(row=row, column=col_idx, value=col_name)
        cell.fill = bleu_greenlog
        cell.font = Font(bold=True, color='FFFFFF')
    
    row += 1
    for _, transp_row in synthese_transp.iterrows():
        ws_synthese[f'A{row}'] = transp_row['Transporteur']
        ws_synthese[f'B{row}'] = f"{transp_row['Total_Indemnisations']:.2f} €"
        ws_synthese[f'C{row}'] = int(transp_row['Nb_Cas'])
        if row % 2 == 0:
            for col_idx in range(1, 4):
                ws_synthese.cell(row=row, column=col_idx).fill = vert_clair
        row += 1
    
    # Ajuster colonnes synthèse
    for col_idx in range(1, 5):
        ws_synthese.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # Sauvegarder
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output

def run():
    """Point d'entrée du module Indemnisations"""
    
    # Initialisation
    if 'indemnisations_data' not in st.session_state:
        st.session_state.indemnisations_data = None
    
    # Chargement automatique
    if st.session_state.indemnisations_data is None:
        saved_data = persistence.load_module_data('indemnisations')
        if saved_data:
            st.session_state.indemnisations_data = saved_data['df']
    
    # Créer DataFrame vide si aucune donnée
    if st.session_state.indemnisations_data is None:
        st.session_state.indemnisations_data = pd.DataFrame(columns=[
            'Date', 'Tracking', 'Partenaire', 'Num_Commande_Origine', 'Num_Commande_Partenaire',
            'Transporteur', 'Motif', 'Montant', 'Statut', 'Notes'
        ])
    
    # En-tête
    col1, col2 = st.columns([4, 1])
    with col1:
        st.title("💶 Module Indemnisations Transporteurs")
        st.markdown("**Suivi des indemnisations reçues par transporteur et par partenaire**")
    with col2:
        if len(st.session_state.indemnisations_data) > 0:
            # Compter les indemnisations "En attente"
            nb_en_attente = len(st.session_state.indemnisations_data[
                st.session_state.indemnisations_data['Statut'] == 'En attente'
            ])
            
            if nb_en_attente > 0:
                st.caption(f"⏳ {nb_en_attente} en attente (conservées)")
            
            if st.button("🗑️ Réinitialiser", type="secondary", key="indem_reset"):
                df = st.session_state.indemnisations_data.copy()
                
                # Filtrer pour ne garder que les "En attente"
                df_en_attente = df[df['Statut'] == 'En attente'].copy()
                
                if len(df_en_attente) > 0:
                    # Sauvegarder uniquement les "En attente"
                    st.session_state.indemnisations_data = df_en_attente
                    persistence.save_module_data('indemnisations', {
                        'df': df_en_attente
                    })
                    nb_supprimees = len(df) - len(df_en_attente)
                    st.success(f"✅ {nb_supprimees} indemnisation(s) supprimée(s) ('{len(df_en_attente)}' En attente conservées)")
                else:
                    # Supprimer tout si aucune "En attente"
                    persistence.delete_module_data('indemnisations')
                    st.session_state.indemnisations_data = pd.DataFrame(columns=[
                        'Date', 'Tracking', 'Partenaire', 'Num_Commande_Origine', 'Num_Commande_Partenaire',
                        'Transporteur', 'Motif', 'Montant', 'Statut', 'Notes'
                    ])
                    st.success("✅ Toutes les indemnisations supprimées")
                
                st.rerun()
    
    st.markdown("---")
    
    # Instructions
    with st.expander("📖 Instructions d'utilisation"):
        st.markdown("""
        ### 📋 Fonctionnalités
        
        Ce module vous permet de :
        - ✅ Enregistrer chaque indemnisation avec **saisie simplifiée** (tracking uniquement)
        - ✅ **Croisement automatique** avec les fichiers logisticien pour trouver le partenaire
        - ✅ Consulter l'historique complet avec toutes les informations
        - ✅ Exporter un récapitulatif mensuel professionnel
        - ✅ Filtrer par période, partenaire ou transporteur
        
        ### 💡 Utilisation Simplifiée
        
        **ÉTAPE PRÉALABLE** : Uploadez vos fichiers logisticien sur la page d'accueil
        
        **Pour ajouter une indemnisation** :
        1. Saisissez uniquement le **numéro de tracking**
        2. Le système recherche **automatiquement** :
           - Le nom du partenaire
           - Le numéro de commande d'origine
           - Le numéro de commande partenaire
        3. Complétez : Date, Transporteur, Motif, Montant, Statut
        4. Cliquez "Ajouter"
        
        ### 🔍 Croisement Automatique
        
        Le système cherche le tracking dans **tous vos fichiers logisticien** :
        - Mois N
        - Mois N-1  
        - Mois N-2
        - Etc.
        
        Si le tracking est trouvé → ✅ Informations récupérées automatiquement  
        Si le tracking n'est pas trouvé → ⚠️ Marqué comme "Non trouvé"
        
        ### 📊 Export Mensuel
        
        En fin de mois :
        1. Filtrez sur le mois concerné
        2. Cliquez "📥 Exporter Récapitulatif"
        3. Obtenez un fichier Excel avec :
           - Synthèse par partenaire
           - Synthèse par transporteur
           - Détail de toutes les indemnisations **avec informations croisées**
        """)
    
    # Statistiques rapides
    if len(st.session_state.indemnisations_data) > 0:
        df = st.session_state.indemnisations_data
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("Total Indemnisations", f"{df['Montant'].sum():.2f} €")
        with col2:
            st.metric("Nombre de cas", len(df))
        with col3:
            st.metric("Partenaires", df['Partenaire'].nunique())
        with col4:
            st.metric("Transporteurs", df['Transporteur'].nunique())
        
        st.markdown("---")
    
    # Onglets
    tab1, tab2, tab3, tab4 = st.tabs(["➕ Ajouter Indemnisation", "📋 Consulter", "✏️ Gérer", "📥 Exporter"])
    
    # TAB 1 : AJOUTER
    with tab1:
        st.subheader("➕ Nouvelle Indemnisation")
        
        # Vérifier si fichiers logisticien disponibles dans la bibliothèque
        from modules.logisticiens_library import load_logisticien_files_for_analysis, get_all_available_periods
        
        log_files = load_logisticien_files_for_analysis(nb_months=6)
        has_log = len(log_files) > 0
        
        if not has_log:
            st.warning("""
            ⚠️ **Aucun fichier logisticien dans la bibliothèque**
            
            Pour utiliser ce module, vous devez d'abord ajouter vos fichiers logisticien dans la bibliothèque.
            
            **Comment faire :**
            1. Allez dans le module **📋 Import Fichier Logisticien**
            2. Uploadez vos fichiers logisticien
            3. Ils seront automatiquement détectés et réutilisés ici
            
            Les fichiers logisticien permettent de croiser automatiquement le tracking avec les informations partenaire.
            """)
        else:
            periods = get_all_available_periods()[:len(log_files)]
            st.success(f"✅ {len(log_files)} fichier(s) logisticien disponible(s) pour croisement automatique")
            with st.expander("📅 Périodes disponibles"):
                from modules.bibliotheque import get_month_name
                for period in periods:
                    st.caption(f"• {get_month_name(period['month'])} {period['year']}")
        
        with st.form("form_indemnisation"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 📋 Informations Requises")
                
                date_indem = st.date_input(
                    "Date de l'indemnisation",
                    value=date.today(),
                    help="Date à laquelle l'indemnisation a été reçue"
                )
                
                tracking = st.text_input(
                    "N° Tracking 🔍",
                    help="Le système recherchera automatiquement le partenaire et les numéros de commande",
                    placeholder="Ex: ABC123456"
                )
                
                # Bouton de recherche manuelle (optionnel, juste pour tester)
                if tracking and has_log:
                    st.caption("🔍 Recherche automatique activée - Le partenaire sera trouvé automatiquement")
            
            with col2:
                st.markdown("#### 💰 Détails Indemnisation")
                
                transporteur = st.selectbox(
                    "Transporteur",
                    options=['DPD', 'Mondial Relay', 'Colissimo', 'Chronopost', 'Colis Privé', 'DHL', 'Autre'],
                    help="Transporteur ayant versé l'indemnisation"
                )
                
                if transporteur == 'Autre':
                    transporteur = st.text_input("Nom du transporteur", key="new_transporteur")
                
                motif = st.selectbox(
                    "Motif",
                    options=[
                        'Colis perdu',
                        'Colis endommagé',
                        'Retard de livraison',
                        'Erreur de facturation',
                        'Non-respect du contrat',
                        'Autre'
                    ],
                    help="Raison de l'indemnisation"
                )
                
                if motif == 'Autre':
                    motif = st.text_input("Précisez le motif", key="motif_autre")
                
                montant = st.number_input(
                    "Montant (€)",
                    min_value=0.0,
                    step=0.01,
                    format="%.2f",
                    help="Montant de l'indemnisation en euros"
                )
                
                statut = st.selectbox(
                    "Statut",
                    options=['Reçue', 'En attente', 'Refusée'],
                    help="Statut de l'indemnisation"
                )
            
            notes = st.text_area(
                "Notes (optionnel)",
                help="Informations complémentaires"
            )
            
            submitted = st.form_submit_button("➕ Ajouter l'Indemnisation", type="primary", use_container_width=True)
            
            if submitted:
                if not tracking:
                    st.error("⚠️ Veuillez saisir un numéro de tracking")
                elif not transporteur or transporteur == 'Autre':
                    st.error("⚠️ Veuillez saisir un transporteur")
                elif montant <= 0:
                    st.error("⚠️ Le montant doit être supérieur à 0")
                else:
                    # Rechercher les informations du tracking
                    with st.spinner("🔍 Recherche des informations dans les fichiers logisticien..."):
                        info = get_info_from_tracking(tracking)
                    
                    # Afficher le résultat de la recherche
                    if info['found']:
                        st.success(f"✅ Tracking trouvé ! Partenaire : **{info['partenaire']}**")
                        if info['num_commande_origine']:
                            st.info(f"📦 Commande origine : {info['num_commande_origine']}")
                        if info['num_commande_partenaire']:
                            st.info(f"🏪 Commande partenaire : {info['num_commande_partenaire']}")
                    else:
                        st.warning(f"⚠️ Tracking non trouvé dans les fichiers logisticien. Partenaire marqué comme '{info['partenaire']}'")
                    
                    # Ajouter la ligne
                    nouvelle_ligne = pd.DataFrame([{
                        'Date': date_indem.strftime('%Y-%m-%d'),
                        'Tracking': tracking,
                        'Partenaire': info['partenaire'],
                        'Num_Commande_Origine': info['num_commande_origine'],
                        'Num_Commande_Partenaire': info['num_commande_partenaire'],
                        'Transporteur': transporteur,
                        'Motif': motif,
                        'Montant': montant,
                        'Statut': statut,
                        'Notes': notes if notes else ''
                    }])
                    
                    st.session_state.indemnisations_data = pd.concat(
                        [st.session_state.indemnisations_data, nouvelle_ligne],
                        ignore_index=True
                    )
                    
                    # Sauvegarder
                    persistence.save_module_data('indemnisations', {
                        'df': st.session_state.indemnisations_data
                    })
                    
                    st.success(f"✅ Indemnisation de {montant:.2f} € ajoutée avec succès !")
                    st.rerun()
    
    # TAB 2 : CONSULTER
    with tab2:
        st.subheader("📋 Historique des Indemnisations")
        
        if len(st.session_state.indemnisations_data) == 0:
            st.info("ℹ️ Aucune indemnisation enregistrée. Ajoutez-en une dans l'onglet 'Ajouter'.")
        else:
            df = st.session_state.indemnisations_data.copy()
            
            # Filtres
            st.markdown("### 🔍 Filtres")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                # Filtre par période
                df['Date'] = pd.to_datetime(df['Date'])
                min_date = df['Date'].min().date()
                max_date = df['Date'].max().date()
                
                date_range = st.date_input(
                    "Période",
                    value=(min_date, max_date),
                    min_value=min_date,
                    max_value=max_date,
                    help="Filtrer par période"
                )
            
            with col2:
                # Filtre par partenaire
                partenaires = ['Tous'] + sorted(df['Partenaire'].unique().tolist())
                partenaire_filter = st.selectbox("Partenaire", partenaires)
            
            with col3:
                # Filtre par transporteur
                transporteurs = ['Tous'] + sorted(df['Transporteur'].unique().tolist())
                transporteur_filter = st.selectbox("Transporteur", transporteurs)
            
            # Appliquer les filtres
            if len(date_range) == 2:
                df = df[(df['Date'] >= pd.to_datetime(date_range[0])) & 
                       (df['Date'] <= pd.to_datetime(date_range[1]))]
            
            if partenaire_filter != 'Tous':
                df = df[df['Partenaire'] == partenaire_filter]
            
            if transporteur_filter != 'Tous':
                df = df[df['Transporteur'] == transporteur_filter]
            
            # Convertir Date en string pour affichage
            df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')
            
            # Afficher statistiques filtrées
            st.markdown("---")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Total période", f"{df['Montant'].sum():.2f} €")
            with col2:
                st.metric("Nombre de cas", len(df))
            with col3:
                nb_recues = len(df[df['Statut'] == 'Reçue'])
                st.metric("Indemnisations reçues", nb_recues)
            
            # Tableau
            st.markdown("### 📊 Détail")
            
            # SECTION RE-MATCHING RÉTROACTIF
            st.markdown("---")
            
            # Vérifier s'il y a des "Non trouvé"
            non_trouves = df[df['Partenaire'] == 'Non trouvé']
            
            if len(non_trouves) > 0:
                with st.expander(f"🔄 Re-Matching Automatique ({len(non_trouves)} indemnisation(s) sans partenaire)", expanded=False):
                    st.markdown("""
                    ### 🎯 Re-Matching Rétroactif
                    
                    Cette fonctionnalité permet de **re-chercher automatiquement** les informations partenaires
                    pour les indemnisations créées **AVANT** l'upload des fichiers logisticiens.
                    
                    **Scénario typique :**
                    1. Vous avez créé des indemnisations → Partenaire "Non trouvé" (fichier logisticien pas encore uploadé)
                    2. Vous uploadez maintenant le fichier logisticien du mois
                    3. Cliquez ci-dessous pour **re-matcher automatiquement** ✅
                    
                    **Le système va :**
                    - ✅ Chercher à nouveau chaque tracking dans les fichiers logisticiens
                    - ✅ Mettre à jour automatiquement les infos si trouvées
                    - ✅ Conserver les indemnisations toujours "Non trouvé" sans modification
                    """)
                    
                    st.info(f"📊 **{len(non_trouves)} indemnisation(s)** avec Partenaire = 'Non trouvé'")
                    
                    # Afficher la liste des trackings "Non trouvé"
                    st.markdown("**Trackings à re-matcher :**")
                    for idx, row in non_trouves.iterrows():
                        st.write(f"• `{row['Tracking']}` - {row['Date']} - {row['Transporteur']}")
                    
                    st.markdown("---")
                    
                    if st.button("🔄 Lancer le Re-Matching", type="primary", use_container_width=True):
                        with st.spinner("🔍 Re-matching en cours..."):
                            # Charger les fichiers logisticiens
                            from modules.logisticiens_library import load_logisticien_files_for_analysis
                            
                            log_files = load_logisticien_files_for_analysis(nb_months=6)
                            
                            if len(log_files) == 0:
                                st.error("""
                                ❌ **Aucun fichier logisticien disponible**
                                
                                Vous devez d'abord uploader des fichiers logisticiens dans le module
                                **"Import Fichier Logisticien"** pour que le re-matching fonctionne.
                                """)
                            else:
                                st.success(f"✅ {len(log_files)} fichier(s) logisticien disponible(s)")
                                
                                # Statistiques
                                nb_updated = 0
                                nb_still_not_found = 0
                                updated_list = []
                                
                                # Re-matcher chaque indemnisation "Non trouvé"
                                df_full = st.session_state.indemnisations_data.copy()
                                
                                for idx, row in df_full.iterrows():
                                    if row['Partenaire'] == 'Non trouvé':
                                        # Re-chercher les infos
                                        info = get_info_from_tracking(row['Tracking'])
                                        
                                        if info['found']:
                                            # Mettre à jour
                                            df_full.at[idx, 'Partenaire'] = info['partenaire']
                                            df_full.at[idx, 'Num_Commande_Origine'] = info['num_commande_origine']
                                            df_full.at[idx, 'Num_Commande_Partenaire'] = info['num_commande_partenaire']
                                            
                                            nb_updated += 1
                                            updated_list.append({
                                                'tracking': row['Tracking'],
                                                'partenaire': info['partenaire']
                                            })
                                        else:
                                            nb_still_not_found += 1
                                
                                # Sauvegarder les modifications
                                if nb_updated > 0:
                                    st.session_state.indemnisations_data = df_full
                                    persistence.save_module_data('indemnisations', {
                                        'df': df_full
                                    })
                                
                                # Afficher le rapport
                                st.markdown("---")
                                st.markdown("### 📊 Rapport du Re-Matching")
                                
                                col1, col2, col3 = st.columns(3)
                                
                                with col1:
                                    st.metric("✅ Mises à jour", nb_updated)
                                
                                with col2:
                                    st.metric("⚠️ Toujours Non trouvé", nb_still_not_found)
                                
                                with col3:
                                    st.metric("📁 Fichiers consultés", len(log_files))
                                
                                if nb_updated > 0:
                                    st.success(f"✅ **{nb_updated} indemnisation(s) mise(s) à jour avec succès !**")
                                    
                                    st.markdown("**Détail des mises à jour :**")
                                    for item in updated_list:
                                        st.write(f"• `{item['tracking']}` → **{item['partenaire']}**")
                                    
                                    st.info("🔄 **Rechargez la page (F5)** pour voir les modifications dans le tableau")
                                
                                if nb_still_not_found > 0:
                                    st.warning(f"""
                                    ⚠️ **{nb_still_not_found} indemnisation(s) toujours "Non trouvé"**
                                    
                                    Ces trackings ne sont pas dans vos fichiers logisticiens uploadés.
                                    
                                    **Solutions possibles :**
                                    - Uploader le fichier logisticien du mois concerné
                                    - Vérifier que le tracking est correct
                                    - Saisir manuellement les informations dans l'onglet "Gérer"
                                    """)
                                
                                if nb_updated == 0 and nb_still_not_found > 0:
                                    st.error("""
                                    ❌ **Aucune correspondance trouvée**
                                    
                                    Les trackings "Non trouvé" ne sont pas dans les fichiers logisticiens uploadés.
                                    Assurez-vous d'avoir uploadé les bons fichiers pour la période concernée.
                                    """)
            
            else:
                st.success("✅ Toutes les indemnisations ont un partenaire identifié !")
            
            st.markdown("---")
            
            # Réorganiser les colonnes pour l'affichage
            display_cols = ['Date', 'Tracking', 'Partenaire', 'Num_Commande_Origine', 'Num_Commande_Partenaire',
                          'Transporteur', 'Motif', 'Montant', 'Statut', 'Notes']
            df_display = df[display_cols]
            
            st.dataframe(
                df_display,
                use_container_width=True,
                hide_index=True,
                column_config={
                    'Date': st.column_config.DateColumn('Date', format='DD/MM/YYYY'),
                    'Tracking': st.column_config.TextColumn('Tracking', width='medium'),
                    'Partenaire': st.column_config.TextColumn('Partenaire', width='medium'),
                    'Num_Commande_Origine': st.column_config.TextColumn('N° Cmd Origine', width='small'),
                    'Num_Commande_Partenaire': st.column_config.TextColumn('N° Cmd Partenaire', width='small'),
                    'Transporteur': st.column_config.TextColumn('Transporteur', width='medium'),
                    'Motif': st.column_config.TextColumn('Motif', width='medium'),
                    'Montant': st.column_config.NumberColumn('Montant', format='%.2f €'),
                    'Statut': st.column_config.TextColumn('Statut', width='small'),
                    'Notes': st.column_config.TextColumn('Notes', width='large'),
                }
            )
    
    # TAB 3 : GÉRER (MODIFIER/SUPPRIMER)
    with tab3:
        st.subheader("✏️ Gérer les Indemnisations")
        
        if len(st.session_state.indemnisations_data) == 0:
            st.info("ℹ️ Aucune indemnisation à gérer. Ajoutez-en une dans l'onglet 'Ajouter'.")
        else:
            df = st.session_state.indemnisations_data.copy()
            
            st.markdown("""
            ### 📊 Tableau de Gestion
            
            **Instructions :**
            - ✏️ **Modifier** : Cliquez sur une cellule pour éditer directement
            - 🗑️ **Supprimer** : Sélectionnez une ligne et cliquez sur le bouton "Supprimer"
            - 💾 Les modifications sont automatiquement sauvegardées
            """)
            
            # Convertir Date en string pour affichage et édition
            df['Date'] = pd.to_datetime(df['Date']).dt.strftime('%Y-%m-%d')
            
            # Configuration des colonnes pour l'éditeur
            column_config = {
                'Date': st.column_config.TextColumn(
                    'Date',
                    help='Date de l\'indemnisation (AAAA-MM-JJ)',
                    width='medium'
                ),
                'Tracking': st.column_config.TextColumn(
                    'Tracking',
                    help='Numéro de tracking',
                    width='medium'
                ),
                'Partenaire': st.column_config.TextColumn(
                    'Partenaire',
                    help='Nom du partenaire (automatique)',
                    width='medium'
                ),
                'Num_Commande_Origine': st.column_config.TextColumn(
                    'Cmd Origine',
                    help='Numéro de commande origine',
                    width='small'
                ),
                'Num_Commande_Partenaire': st.column_config.TextColumn(
                    'Cmd Partenaire',
                    help='Numéro de commande partenaire',
                    width='small'
                ),
                'Transporteur': st.column_config.SelectboxColumn(
                    'Transporteur',
                    help='Transporteur',
                    width='medium',
                    options=['DPD', 'Mondial Relay', 'Colissimo', 'Chronopost', 'Colis Privé', 'DHL', 'Autre'],
                    required=True
                ),
                'Motif': st.column_config.SelectboxColumn(
                    'Motif',
                    help='Motif de l\'indemnisation',
                    width='medium',
                    options=[
                        'Colis perdu',
                        'Colis endommagé',
                        'Retard de livraison',
                        'Erreur de facturation',
                        'Non-respect du contrat',
                        'Autre'
                    ],
                    required=True
                ),
                'Montant': st.column_config.NumberColumn(
                    'Montant (€)',
                    help='Montant de l\'indemnisation',
                    format='%.2f',
                    min_value=0.0,
                    step=0.01
                ),
                'Statut': st.column_config.SelectboxColumn(
                    'Statut',
                    help='Statut de l\'indemnisation',
                    width='small',
                    options=['Reçue', 'En attente', 'Refusée'],
                    required=True
                ),
                'Notes': st.column_config.TextColumn(
                    'Notes',
                    help='Notes complémentaires',
                    width='large'
                )
            }
            
            # Éditeur de données avec sélection multi-lignes
            edited_result = st.data_editor(
                df,
                use_container_width=True,
                hide_index=False,
                column_config=column_config,
                num_rows="fixed",
                key="indemnisations_editor",
                on_change=None,
                disabled=False
            )
            
            # Boutons d'action
            st.markdown("---")
            st.markdown("### ⚙️ Actions")
            
            col1, col2, col3 = st.columns([2, 1, 1])
            
            with col1:
                st.markdown("""
                **Instructions :**
                - ✏️ Cliquez sur une cellule pour modifier
                - 🗑️ Pour supprimer : entrez les numéros de lignes séparés par des virgules (ex: 0,2,5)
                """)
            
            with col2:
                if st.button("💾 Sauvegarder Modifications", type="primary", use_container_width=True):
                    # Mettre à jour les données
                    st.session_state.indemnisations_data = edited_result.copy()
                    
                    # Sauvegarder
                    persistence.save_module_data('indemnisations', {
                        'df': st.session_state.indemnisations_data
                    })
                    
                    st.success("✅ Modifications sauvegardées avec succès !")
                    st.rerun()
            
            with col3:
                # Input pour les lignes à supprimer
                lignes_a_supprimer = st.text_input(
                    "Lignes à supprimer",
                    placeholder="Ex: 0,2,5",
                    help="Entrez les numéros de lignes séparés par des virgules (première ligne = 0)",
                    key="delete_lines_input"
                )
            
            # Bouton de suppression
            if lignes_a_supprimer:
                try:
                    # Parser les indices
                    indices = [int(x.strip()) for x in lignes_a_supprimer.split(',') if x.strip()]
                    # Filtrer les indices valides
                    indices_valides = [i for i in indices if 0 <= i < len(df)]
                    
                    if indices_valides:
                        st.warning(f"⚠️ Vous allez supprimer {len(indices_valides)} ligne(s) : {', '.join([str(i) for i in indices_valides])}")
                        
                        col_confirm1, col_confirm2, col_confirm3 = st.columns([2, 1, 1])
                        with col_confirm2:
                            if st.button("✅ Confirmer Suppression", type="secondary", use_container_width=True):
                                # Supprimer les lignes
                                st.session_state.indemnisations_data = df.drop(indices_valides).reset_index(drop=True)
                                
                                # Sauvegarder
                                persistence.save_module_data('indemnisations', {
                                    'df': st.session_state.indemnisations_data
                                })
                                
                                st.success(f"✅ {len(indices_valides)} ligne(s) supprimée(s) !")
                                st.rerun()
                        
                        with col_confirm3:
                            if st.button("❌ Annuler", use_container_width=True):
                                st.rerun()
                    else:
                        st.error("❌ Aucun indice valide trouvé")
                except ValueError:
                    st.error("❌ Format invalide. Utilisez des nombres séparés par des virgules (ex: 0,2,5)")
            
            # Statistiques après modifications
            st.markdown("---")
            st.markdown("### 📊 Statistiques")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total", f"{edited_result['Montant'].sum():.2f} €")
            with col2:
                st.metric("Nombre", len(edited_result))
            with col3:
                nb_recues = len(edited_result[edited_result['Statut'] == 'Reçue'])
                st.metric("Reçues", nb_recues)
            with col4:
                nb_attente = len(edited_result[edited_result['Statut'] == 'En attente'])
                st.metric("En attente", nb_attente)
    
    # TAB 4 : EXPORTER
    with tab4:
        st.subheader("📥 Exporter Récapitulatif")
        
        if len(st.session_state.indemnisations_data) == 0:
            st.info("ℹ️ Aucune indemnisation à exporter.")
        else:
            df = st.session_state.indemnisations_data.copy()
            
            st.markdown("""
            ### 📅 Période d'Export
            
            Sélectionnez la période pour laquelle vous souhaitez générer le récapitulatif :
            """)
            
            # Sélection période
            df['Date'] = pd.to_datetime(df['Date'])
            min_date = df['Date'].min().date()
            max_date = df['Date'].max().date()
            
            col1, col2 = st.columns(2)
            with col1:
                date_debut = st.date_input(
                    "Date de début",
                    value=min_date,
                    min_value=min_date,
                    max_value=max_date
                )
            with col2:
                date_fin = st.date_input(
                    "Date de fin",
                    value=max_date,
                    min_value=min_date,
                    max_value=max_date
                )
            
            # Filtrer
            df_export = df[(df['Date'] >= pd.to_datetime(date_debut)) & 
                          (df['Date'] <= pd.to_datetime(date_fin))].copy()
            
            # Convertir Date en string
            df_export['Date'] = df_export['Date'].dt.strftime('%Y-%m-%d')
            
            # Statistiques
            st.markdown("---")
            st.markdown("### 📊 Aperçu de l'Export")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Période", f"{(pd.to_datetime(date_fin) - pd.to_datetime(date_debut)).days + 1} jours")
            with col2:
                st.metric("Total", f"{df_export['Montant'].sum():.2f} €")
            with col3:
                st.metric("Cas", len(df_export))
            with col4:
                st.metric("Partenaires", df_export['Partenaire'].nunique())
            
            # Synthèse par transporteur
            st.markdown("#### Par Transporteur")
            synthese_transp = df_export.groupby('Transporteur')['Montant'].sum().sort_values(ascending=False)
            cols = st.columns(len(synthese_transp))
            for idx, (transp, montant) in enumerate(synthese_transp.items()):
                with cols[idx]:
                    st.metric(transp, f"{montant:.2f} €")
            
            # Bouton export
            st.markdown("---")
            
            if len(df_export) == 0:
                st.warning("⚠️ Aucune indemnisation sur cette période")
            else:
                if st.button("📥 Générer Export Excel", type="primary", use_container_width=True):
                    with st.spinner("Génération du fichier Excel..."):
                        excel_file = export_indemnisations_excel(df_export)
                        
                        filename = f"Indemnisations_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}.xlsx"
                        
                        st.success("✅ Export généré avec succès !")
                        
                        st.download_button(
                            label=f"📥 Télécharger {filename}",
                            data=excel_file,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True
                        )
