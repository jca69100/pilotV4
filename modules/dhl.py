import streamlit as st
import pandas as pd
import csv
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from shared import persistence

def process_dhl_file(uploaded_file):
    """Traite le fichier DHL et génère les données de facturation"""
    
    # Lire le fichier CSV
    content = uploaded_file.read().decode('utf-8')
    csv_reader = csv.DictReader(content.splitlines())
    
    shipments = []
    
    for row in csv_reader:
        # Ne garder que les lignes de type "S" (Shipment)
        if row['Line Type'] == 'S':
            shipment = {
                'Numero_Expedition': row['Shipment Number'],
                'Date_Expedition': row['Shipment Date'],
                'Expediteur': row['Senders Name'],
                'Destination': f"{row['Dest Name']} - {row['Dest Country Code']}",
                'Poids_kg': row['Weight (kg)'],
                'Devise': row['Currency'],
                
                # Tarif de base (Colonne G)
                'Tarif_Base_HT': row['Weight Charge'],
                'Tarif_Base_Taxe': row['Weight Tax (VAT)'],
                
                # XC1 (Colonne K)
                'XC1_Code': row['XC1 Code'],
                'XC1_Nom': row['XC1 Name'],
                'XC1_Montant_HT': row['XC1 Charge'],
                'XC1_Taxe': row['XC1 Tax'],
                'XC1_Total_TTC': row['XC1 Total'],
                
                # XC2 (Colonne P)
                'XC2_Code': row['XC2 Code'],
                'XC2_Nom': row['XC2 Name'],
                'XC2_Montant_HT': row['XC2 Charge'],
                'XC2_Taxe': row['XC2 Tax'],
                'XC2_Total_TTC': row['XC2 Total'],
                
                # XC3 (Colonne U)
                'XC3_Code': row['XC3 Code'],
                'XC3_Nom': row['XC3 Name'],
                'XC3_Montant_HT': row['XC3 Charge'],
                'XC3_Taxe': row['XC3 Tax'],
                'XC3_Total_TTC': row['XC3 Total'],
                
                # XC4 (Colonne Z)
                'XC4_Code': row['XC4 Code'],
                'XC4_Nom': row['XC4 Name'],
                'XC4_Montant_HT': row['XC4 Charge'],
                'XC4_Taxe': row['XC4 Tax'],
                'XC4_Total_TTC': row['XC4 Total'],
                
                # XC5
                'XC5_Code': row['XC5 Code'],
                'XC5_Nom': row['XC5 Name'],
                'XC5_Montant_HT': row['XC5 Charge'],
                'XC5_Taxe': row['XC5 Tax'],
                'XC5_Total_TTC': row['XC5 Total'],
                
                # Totaux (Colonne AL)
                'Total_Frais_Supp_HT': row['Total Extra Charges (XC)'],
                'Total_Frais_Supp_Taxe': row['Total Extra Charges Tax'],
                'Total_HT': row['Total amount (excl. VAT)'],
                'Total_Taxes': row['Total Tax'],
                'Total_TTC': row['Total amount (incl. VAT)']
            }
            shipments.append(shipment)
    
    return pd.DataFrame(shipments)


def enrich_with_logisticiens(df):
    """Enrichit les données DHL avec les informations des fichiers logisticiens"""
    
    # Charger les fichiers logisticiens depuis la bibliothèque
    from modules.logisticiens_library import load_logisticien_files_for_analysis
    
    log_files = load_logisticien_files_for_analysis(nb_months=3)
    
    if len(log_files) == 0:
        st.warning("⚠️ Aucun fichier logisticien disponible - Enrichissement impossible")
        # Ajouter colonnes vides
        df['Partenaire'] = 'Non trouvé'
        df['Num_Commande_Origine'] = ''
        df['Num_Commande_Partenaire'] = ''
        df['Match'] = False
        return df
    
    # Fusionner tous les fichiers logisticiens
    all_log_data = []
    for log_file in log_files:
        try:
            log_file.seek(0)
            df_log_temp = pd.read_excel(log_file, sheet_name='Facturation préparation')
            all_log_data.append(df_log_temp)
        except Exception as e:
            st.warning(f"⚠️ Erreur lecture fichier: {str(e)}")
    
    if not all_log_data:
        df['Partenaire'] = 'Non trouvé'
        df['Num_Commande_Origine'] = ''
        df['Num_Commande_Partenaire'] = ''
        df['Match'] = False
        return df
    
    df_log = pd.concat(all_log_data, ignore_index=True)
    df_log = df_log.drop_duplicates(subset=['Numéro de tracking'], keep='first')
    
    # Nettoyer les trackings
    df['Tracking_Clean'] = df['Numero_Expedition'].astype(str).str.strip().str.replace('.0', '', regex=False)
    df_log['Tracking_Clean'] = df_log['Numéro de tracking'].astype(str).str.strip().str.replace('.0', '', regex=False)
    
    # Merger
    df_merged = df.merge(
        df_log[['Tracking_Clean', 'Nom du partenaire', "Numéro de commande d'origine", 
                'Numéro de commande partenaire']],
        on='Tracking_Clean',
        how='left'
    )
    
    # Renommer colonnes
    df_merged['Partenaire'] = df_merged['Nom du partenaire'].fillna('Non trouvé')
    df_merged['Num_Commande_Origine'] = df_merged["Numéro de commande d'origine"].fillna('')
    df_merged['Num_Commande_Partenaire'] = df_merged['Numéro de commande partenaire'].fillna('')
    df_merged['Match'] = df_merged['Nom du partenaire'].notna()
    
    # Supprimer colonnes temporaires
    df_merged = df_merged.drop(['Tracking_Clean', 'Nom du partenaire', 
                                 "Numéro de commande d'origine", 
                                 'Numéro de commande partenaire'], axis=1)
    
    nb_matched = df_merged['Match'].sum()
    match_rate = (nb_matched / len(df_merged) * 100) if len(df_merged) > 0 else 0
    
    st.info(f"📊 Matching : {nb_matched}/{len(df_merged)} trackings trouvés ({match_rate:.1f}%)")
    
    return df_merged


def create_synthese_colonnes(df):
    """Crée le tableau de synthèse des colonnes G, K, P, U, Z, AL"""
    
    # Convertir en numérique
    numeric_cols = ['Tarif_Base_HT', 'XC1_Montant_HT', 'XC2_Montant_HT', 
                    'XC3_Montant_HT', 'XC4_Montant_HT', 'Total_TTC']
    
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    synthese_data = []
    
    # Colonne G - Tarif_Base_HT
    synthese_data.append({
        'Colonne Excel': 'G',
        'Nom Colonne': 'Tarif_Base_HT',
        'Description': 'Tarif de base hors taxes (Weight Charge)',
        'Nb Lignes': len(df),
        'Montant Total (€)': df['Tarif_Base_HT'].sum(),
        'Montant Moyen (€)': df['Tarif_Base_HT'].mean(),
        'Montant Min (€)': df['Tarif_Base_HT'].min(),
        'Montant Max (€)': df['Tarif_Base_HT'].max()
    })
    
    # Colonne K - XC1_Montant_HT (généralement Fuel Surcharge)
    xc1_types = df[df['XC1_Montant_HT'] > 0]['XC1_Nom'].value_counts()
    xc1_desc = xc1_types.index[0] if len(xc1_types) > 0 else 'N/A'
    
    synthese_data.append({
        'Colonne Excel': 'K',
        'Nom Colonne': 'XC1_Montant_HT',
        'Description': f'Frais supplémentaire 1 ({xc1_desc})',
        'Nb Lignes': len(df[df['XC1_Montant_HT'] > 0]),
        'Montant Total (€)': df['XC1_Montant_HT'].sum(),
        'Montant Moyen (€)': df[df['XC1_Montant_HT'] > 0]['XC1_Montant_HT'].mean() if len(df[df['XC1_Montant_HT'] > 0]) > 0 else 0,
        'Montant Min (€)': df[df['XC1_Montant_HT'] > 0]['XC1_Montant_HT'].min() if len(df[df['XC1_Montant_HT'] > 0]) > 0 else 0,
        'Montant Max (€)': df['XC1_Montant_HT'].max()
    })
    
    # Colonne P - XC2_Montant_HT (généralement Demand Surcharge)
    xc2_types = df[df['XC2_Montant_HT'] > 0]['XC2_Nom'].value_counts()
    xc2_desc = xc2_types.index[0] if len(xc2_types) > 0 else 'N/A'
    
    synthese_data.append({
        'Colonne Excel': 'P',
        'Nom Colonne': 'XC2_Montant_HT',
        'Description': f'Frais supplémentaire 2 ({xc2_desc})',
        'Nb Lignes': len(df[df['XC2_Montant_HT'] > 0]),
        'Montant Total (€)': df['XC2_Montant_HT'].sum(),
        'Montant Moyen (€)': df[df['XC2_Montant_HT'] > 0]['XC2_Montant_HT'].mean() if len(df[df['XC2_Montant_HT'] > 0]) > 0 else 0,
        'Montant Min (€)': df[df['XC2_Montant_HT'] > 0]['XC2_Montant_HT'].min() if len(df[df['XC2_Montant_HT'] > 0]) > 0 else 0,
        'Montant Max (€)': df['XC2_Montant_HT'].max()
    })
    
    # Colonne U - XC3_Montant_HT (généralement GoGreen/Remote Area)
    xc3_types = df[df['XC3_Montant_HT'] > 0]['XC3_Nom'].value_counts()
    xc3_desc = xc3_types.index[0] if len(xc3_types) > 0 else 'N/A'
    
    synthese_data.append({
        'Colonne Excel': 'U',
        'Nom Colonne': 'XC3_Montant_HT',
        'Description': f'Frais supplémentaire 3 ({xc3_desc})',
        'Nb Lignes': len(df[df['XC3_Montant_HT'] > 0]),
        'Montant Total (€)': df['XC3_Montant_HT'].sum(),
        'Montant Moyen (€)': df[df['XC3_Montant_HT'] > 0]['XC3_Montant_HT'].mean() if len(df[df['XC3_Montant_HT'] > 0]) > 0 else 0,
        'Montant Min (€)': df[df['XC3_Montant_HT'] > 0]['XC3_Montant_HT'].min() if len(df[df['XC3_Montant_HT'] > 0]) > 0 else 0,
        'Montant Max (€)': df['XC3_Montant_HT'].max()
    })
    
    # Colonne Z - XC4_Montant_HT
    xc4_types = df[df['XC4_Montant_HT'] > 0]['XC4_Nom'].value_counts()
    xc4_desc = xc4_types.index[0] if len(xc4_types) > 0 else 'N/A'
    
    synthese_data.append({
        'Colonne Excel': 'Z',
        'Nom Colonne': 'XC4_Montant_HT',
        'Description': f'Frais supplémentaire 4 ({xc4_desc})',
        'Nb Lignes': len(df[df['XC4_Montant_HT'] > 0]),
        'Montant Total (€)': df['XC4_Montant_HT'].sum(),
        'Montant Moyen (€)': df[df['XC4_Montant_HT'] > 0]['XC4_Montant_HT'].mean() if len(df[df['XC4_Montant_HT'] > 0]) > 0 else 0,
        'Montant Min (€)': df[df['XC4_Montant_HT'] > 0]['XC4_Montant_HT'].min() if len(df[df['XC4_Montant_HT'] > 0]) > 0 else 0,
        'Montant Max (€)': df['XC4_Montant_HT'].max()
    })
    
    # Colonne AP - Total_TTC
    synthese_data.append({
        'Colonne Excel': 'AP',
        'Nom Colonne': 'Total_TTC',
        'Description': 'Montant total TTC (Total amount incl. VAT)',
        'Nb Lignes': len(df),
        'Montant Total (€)': df['Total_TTC'].sum(),
        'Montant Moyen (€)': df['Total_TTC'].mean(),
        'Montant Min (€)': df['Total_TTC'].min(),
        'Montant Max (€)': df['Total_TTC'].max()
    })
    
    df_synthese = pd.DataFrame(synthese_data)
    
    # Arrondir
    for col in ['Montant Total (€)', 'Montant Moyen (€)', 'Montant Min (€)', 'Montant Max (€)']:
        df_synthese[col] = df_synthese[col].round(2)
    
    return df_synthese


def export_excel_dhl(df, synthese_colonnes):
    """Génère le fichier Excel avec mise en forme et colonnes en évidence"""
    
    output = BytesIO()
    wb = Workbook()
    
    # Feuille 1: Données détaillées
    ws1 = wb.active
    ws1.title = "Données Détaillées"
    
    # Réorganiser les colonnes : numéros de commande en premier
    cols_order = [
        'Num_Commande_Origine', 'Num_Commande_Partenaire', 'Partenaire', 'Match',
        'Numero_Expedition', 'Date_Expedition', 'Expediteur', 'Destination', 'Poids_kg', 'Devise',
        'Tarif_Base_HT', 'Tarif_Base_Taxe',
        'XC1_Code', 'XC1_Nom', 'XC1_Montant_HT', 'XC1_Taxe', 'XC1_Total_TTC',
        'XC2_Code', 'XC2_Nom', 'XC2_Montant_HT', 'XC2_Taxe', 'XC2_Total_TTC',
        'XC3_Code', 'XC3_Nom', 'XC3_Montant_HT', 'XC3_Taxe', 'XC3_Total_TTC',
        'XC4_Code', 'XC4_Nom', 'XC4_Montant_HT', 'XC4_Taxe', 'XC4_Total_TTC',
        'XC5_Code', 'XC5_Nom', 'XC5_Montant_HT', 'XC5_Taxe', 'XC5_Total_TTC',
        'Total_Frais_Supp_HT', 'Total_Frais_Supp_Taxe',
        'Total_HT', 'Total_Taxes', 'Total_TTC'
    ]
    
    df_export = df[cols_order]
    
    # Écrire le DataFrame
    for r in dataframe_to_rows(df_export, index=False, header=True):
        ws1.append(r)
    
    # Styles
    bleu_header = PatternFill(start_color='2D3E50', end_color='2D3E50', fill_type='solid')
    jaune_highlight = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    font_header = Font(bold=True, color='FFFFFF', size=11)
    font_normal = Font(size=10)
    
    # En-tête
    for cell in ws1[1]:
        cell.fill = bleu_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Colonnes à mettre en évidence (positions après réorganisation)
    # Tarif_Base_HT est colonne 11 (K)
    # XC1_Montant_HT est colonne 15 (O) 
    # XC2_Montant_HT est colonne 20 (T)
    # XC3_Montant_HT est colonne 25 (Y)
    # XC4_Montant_HT est colonne 30 (AD)
    # Total_TTC est colonne 42 (AP)
    
    highlight_cols = [11, 15, 20, 25, 30, 42]  # Index 1-based
    
    for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
        for idx, cell in enumerate(row, start=1):
            if idx in highlight_cols:
                cell.fill = jaune_highlight
            cell.font = font_normal
            cell.alignment = Alignment(horizontal='left' if idx <= 10 else 'right')
    
    # Largeurs de colonnes
    ws1.column_dimensions['A'].width = 18  # Num Commande Origine
    ws1.column_dimensions['B'].width = 18  # Num Commande Partenaire
    ws1.column_dimensions['C'].width = 20  # Partenaire
    ws1.column_dimensions['D'].width = 8   # Match
    ws1.column_dimensions['E'].width = 18  # Numero Expedition
    
    # Figer la première ligne
    ws1.freeze_panes = 'A2'
    
    # Feuille 2: Synthèse colonnes
    ws2 = wb.create_sheet("Synthèse Colonnes")
    
    for r in dataframe_to_rows(synthese_colonnes, index=False, header=True):
        ws2.append(r)
    
    # En-tête
    for cell in ws2[1]:
        cell.fill = bleu_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Ajuster largeurs
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 45
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 18
    ws2.column_dimensions['G'].width = 18
    ws2.column_dimensions['H'].width = 18
    
    # Figer
    ws2.freeze_panes = 'A2'
    
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def run():
    """Module principal DHL"""
    
    st.title("📦 DHL - Facturation")
    
    # Initialiser session state
    if 'dhl_data' not in st.session_state:
        st.session_state.dhl_data = None
    if 'dhl_files' not in st.session_state:
        st.session_state.dhl_files = {'facture': None}
    
    # Charger données sauvegardées
    if st.session_state.dhl_data is None:
        saved_data = persistence.load_module_data('dhl')
        if saved_data:
            st.session_state.dhl_data = saved_data
            st.success("✅ Données DHL chargées depuis la sauvegarde")
    
    # Vérifier fichiers logisticiens
    from modules.logisticiens_library import load_logisticien_files_for_analysis, get_all_available_periods
    
    log_files = load_logisticien_files_for_analysis(nb_months=3)
    
    st.markdown("---")
    st.markdown("### 📋 Fichiers Logisticien")
    
    if len(log_files) > 0:
        periods = get_all_available_periods()[:len(log_files)]
        st.success(f"✅ {len(log_files)} fichier(s) logisticien disponible(s)")
        with st.expander("📅 Périodes disponibles"):
            from modules.bibliotheque import get_month_name
            for period in periods:
                st.caption(f"• {get_month_name(period['month'])} {period['year']}")
    else:
        st.warning("⚠️ Aucun fichier logisticien dans la bibliothèque")
        st.caption("Le croisement avec les partenaires ne sera pas possible")
    
    st.markdown("---")
    
    # Afficher données ou upload
    if st.session_state.dhl_data is not None:
        # Données déjà traitées
        df = st.session_state.dhl_data['df']
        synthese_colonnes = st.session_state.dhl_data['synthese_colonnes']
        
        # Statistiques globales
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Total expéditions", len(df))
        with col2:
            total_ttc = pd.to_numeric(df['Total_TTC'], errors='coerce').sum()
            st.metric("Total TTC", f"{total_ttc:,.2f} €")
        with col3:
            nb_matched = df['Match'].sum() if 'Match' in df.columns else 0
            st.metric("Trackings matchés", f"{nb_matched}/{len(df)}")
        with col4:
            nb_partners = df['Partenaire'].nunique() if 'Partenaire' in df.columns else 0
            st.metric("Partenaires", nb_partners)
        
        st.markdown("---")
        
        # Onglets
        tab1, tab2, tab3 = st.tabs(["📊 Synthèse Colonnes", "📋 Données Détaillées", "📥 Export"])
        
        with tab1:
            st.subheader("📊 Synthèse des Colonnes Clés")
            st.markdown("""
            Tableau détaillant les colonnes mises en évidence dans l'export Excel :
            - **Colonne G** : Tarif de base HT
            - **Colonne K** : Frais supplémentaire 1 (généralement Fuel Surcharge)
            - **Colonne P** : Frais supplémentaire 2 (généralement Demand Surcharge)
            - **Colonne U** : Frais supplémentaire 3 (généralement GoGreen / Remote Area)
            - **Colonne Z** : Frais supplémentaire 4
            - **Colonne AP** : Total TTC
            """)
            
            st.dataframe(synthese_colonnes, use_container_width=True, hide_index=True)
        
        with tab2:
            st.subheader("📋 Données Détaillées")
            
            # Filtres
            col1, col2 = st.columns(2)
            with col1:
                partenaires = ['Tous'] + sorted(df['Partenaire'].unique().tolist())
                selected_partner = st.selectbox("Filtrer par partenaire", partenaires)
            with col2:
                show_only_matched = st.checkbox("Afficher uniquement les trackings matchés", value=False)
            
            # Filtrer
            df_display = df.copy()
            if selected_partner != 'Tous':
                df_display = df_display[df_display['Partenaire'] == selected_partner]
            if show_only_matched:
                df_display = df_display[df_display['Match'] == True]
            
            # Réorganiser colonnes : numéros de commande en premier
            cols_display = ['Num_Commande_Origine', 'Num_Commande_Partenaire', 'Partenaire', 'Match']
            cols_display += [col for col in df_display.columns if col not in cols_display]
            df_display = df_display[cols_display]
            
            st.dataframe(df_display, use_container_width=True, hide_index=True)
            
            st.caption(f"📊 {len(df_display)} expéditions affichées sur {len(df)} total")
        
        with tab3:
            st.subheader("📥 Export Excel")
            
            st.markdown("""
            Le fichier Excel contient :
            - **Feuille 1** : Données détaillées avec colonnes G, K, P, U, Z, AP en surbrillance jaune
            - **Feuille 2** : Synthèse des colonnes clés
            
            Les numéros de commande sont placés en colonnes A et B.
            """)
            
            excel_data = export_excel_dhl(df, synthese_colonnes)
            
            st.download_button(
                label="📥 Télécharger Excel",
                data=excel_data,
                file_name=f"DHL_Facturation_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
    else:
        # Upload de fichier
        st.subheader("📤 Upload Facture DHL")
        
        uploaded_file = st.file_uploader(
            "Sélectionnez le fichier DHL (DHLFR_TD_concatenated_*.csv)",
            type=['csv'],
            help="Format: DHLFR_TD_concatenated_YYYYMMDD_HHMMSS.csv"
        )
        
        if uploaded_file:
            with st.spinner("Traitement du fichier DHL..."):
                try:
                    # Traiter le fichier
                    df = process_dhl_file(uploaded_file)
                    
                    if len(df) == 0:
                        st.error("❌ Aucune expédition trouvée dans le fichier")
                        st.stop()
                    
                    st.success(f"✅ {len(df)} expéditions extraites")
                    
                    # Enrichir avec logisticiens
                    df = enrich_with_logisticiens(df)
                    
                    # Créer synthèse colonnes
                    synthese_colonnes = create_synthese_colonnes(df)
                    
                    # Sauvegarder
                    st.session_state.dhl_data = {
                        'df': df,
                        'synthese_colonnes': synthese_colonnes,
                        'timestamp': datetime.now()
                    }
                    st.session_state.dhl_files['facture'] = uploaded_file.name
                    
                    # Sauvegarde persistante
                    persistence.save_module_files('dhl', st.session_state.dhl_files)
                    persistence.save_module_data('dhl', st.session_state.dhl_data)
                    
                    # 📚 AUTO-ARCHIVAGE DANS LA BIBLIOTHÈQUE
                    success, year, month = persistence.auto_archive_analysis(
                        'DHL',
                        df,
                        st.session_state.dhl_data
                    )
                    
                    # Message avec info archivage
                    if success:
                        from modules.bibliotheque import get_month_name
                        st.success(f"✅ Analyse terminée et archivée ({get_month_name(month)} {year})")
                    else:
                        st.success("✅ Analyse terminée et sauvegardée !")
                    
                    st.rerun()
                    
                except Exception as e:
                    st.error(f"❌ Erreur lors du traitement : {str(e)}")
                    import traceback
                    st.code(traceback.format_exc())
