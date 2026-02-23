"""
Module Export Global Multi-Transporteurs pour GREENLOG
Export consolidé par COMMANDE avec colonnes par transporteur
Version 2.0 - Vue orientée commande
"""

import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_consolidated_export(partenaires_filter=None):
    """
    Créer un export Excel simplifié pour REFACTURATION
    Focus : Coûts à refacturer par partenaire et commande
    
    Args:
        partenaires_filter: Liste des partenaires à inclure (None = tous)
    
    Returns:
        BytesIO: Fichier Excel
    """
    
    # Dictionnaire pour stocker les coûts par tracking
    all_data = {}
    
    # Couleurs GREENLOG
    bleu_greenlog = PatternFill(start_color='2D3E50', end_color='2D3E50', fill_type='solid')
    rouge_alert = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    vert_clair = PatternFill(start_color='E8F5F1', end_color='E8F5F1', fill_type='solid')
    jaune_alert = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
    
    # ==================== COLLECTER COÛTS DPD ====================
    if 'dpd_data' in st.session_state and st.session_state.dpd_data:
        try:
            df_dpd = st.session_state.dpd_data['df_with_taxes'].copy()
            
            for _, row in df_dpd.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Partenaire': partner,
                        'Tracking': tracking,
                        'Num_Commande': row.get("Numéro de commande d'origine", ''),
                        'Date': row.get('Date de la commande', ''),
                    }
                
                # Coût DPD = Total TTC
                all_data[tracking]['DPD'] = row.get('Montant total TTC', 0)
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte DPD: {str(e)}")
    
    # ==================== COLLECTER COÛTS MONDIAL RELAY ====================
    if 'mondial_relay_data' in st.session_state and st.session_state.mondial_relay_data:
        try:
            df_mr = st.session_state.mondial_relay_data['df'].copy()
            
            for _, row in df_mr.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Partenaire': partner,
                        'Tracking': tracking,
                        'Num_Commande': row.get("Numéro de commande d'origine", ''),
                        'Date': row.get('Date de la commande', ''),
                    }
                
                # Coût Mondial Relay = Total avec taxe
                all_data[tracking]['Mondial_Relay'] = row.get('Total avec taxe', 0)
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Mondial Relay: {str(e)}")
    
    # ==================== COLLECTER COÛTS COLISSIMO ====================
    # Note: Colissimo n'a pas de coût direct, juste statut
    if 'colissimo_data' in st.session_state and st.session_state.colissimo_data:
        try:
            df_col = st.session_state.colissimo_data['df_matched'].copy()
            
            for _, row in df_col.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Partenaire': partner,
                        'Tracking': tracking,
                        'Num_Commande': row.get("Numéro de commande d'origine", ''),
                        'Date': row.get('Date de la commande', ''),
                    }
                
                # Colissimo : pas de coût direct, juste indication
                all_data[tracking]['Colissimo'] = 0
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Colissimo: {str(e)}")
    
    # ==================== COLLECTER COÛTS CHRONOPOST ====================
    if 'chronopost_data' in st.session_state and st.session_state.chronopost_data:
        try:
            df_chrono = st.session_state.chronopost_data['df'].copy()
            
            for _, row in df_chrono.iterrows():
                tracking = str(row.get('Tracking', ''))
                partner = row.get('Partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Partenaire': partner,
                        'Tracking': tracking,
                        'Num_Commande': row.get('Num_Commande_Origine', ''),
                        'Date': row.get('Date', ''),
                    }
                
                # Coût Chronopost = Écart prix (si positif) + Surplus
                ecart = row.get('Difference_Prix', 0)
                all_data[tracking]['Chronopost'] = ecart if ecart > 0 else 0
            
            # Ajouter les surplus
            if 'df_surplus' in st.session_state.chronopost_data:
                df_surplus = st.session_state.chronopost_data['df_surplus'].copy()
                
                surplus_by_tracking = df_surplus.groupby('Tracking')['Montant_Surplus'].sum().to_dict()
                
                for tracking, surplus in surplus_by_tracking.items():
                    tracking_str = str(tracking)
                    if tracking_str in all_data:
                        current = all_data[tracking_str].get('Chronopost', 0)
                        all_data[tracking_str]['Chronopost'] = current + surplus
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Chronopost: {str(e)}")
    
    # ==================== COLLECTER COÛTS COLIS PRIVÉ ====================
    if 'colis_prive_data' in st.session_state and st.session_state.colis_prive_data:
        try:
            df_cp = st.session_state.colis_prive_data['df'].copy()
            
            for _, row in df_cp.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Partenaire': partner,
                        'Tracking': tracking,
                        'Num_Commande': row.get("Numéro de commande d'origine", ''),
                        'Date': row.get('Date de la commande', ''),
                    }
                
                # Coût Colis Privé = Majoration
                all_data[tracking]['Colis_Prive'] = row.get('Majoration service', 0)
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Colis Privé: {str(e)}")
    
    # ==================== CRÉER LE DATAFRAME SIMPLIFIÉ ====================
    if not all_data:
        return None
    
    df_refacturation = pd.DataFrame.from_dict(all_data, orient='index')
    
    # Colonnes dans l'ordre voulu
    colonnes_ordre = [
        'Partenaire',
        'Num_Commande', 
        'Tracking',
        'Date',
        'DPD',
        'Mondial_Relay',
        'Colissimo',
        'Chronopost',
        'Colis_Prive'
    ]
    
    # Ajouter colonnes manquantes avec 0
    for col in colonnes_ordre:
        if col not in df_refacturation.columns:
            if col in ['DPD', 'Mondial_Relay', 'Colissimo', 'Chronopost', 'Colis_Prive']:
                df_refacturation[col] = 0
            else:
                df_refacturation[col] = ''
    
    df_refacturation = df_refacturation[colonnes_ordre]
    
    # Remplacer NaN par 0 pour les colonnes numériques
    cols_numeriques = ['DPD', 'Mondial_Relay', 'Colissimo', 'Chronopost', 'Colis_Prive']
    for col in cols_numeriques:
        df_refacturation[col] = df_refacturation[col].fillna(0)
    
    # Calculer le TOTAL par ligne
    df_refacturation['TOTAL_A_REFACTURER'] = df_refacturation[cols_numeriques].sum(axis=1)
    
    # Supprimer les lignes où le total = 0 (pas de coûts à refacturer)
    df_refacturation = df_refacturation[df_refacturation['TOTAL_A_REFACTURER'] > 0]
    
    # Trier par partenaire puis par total décroissant
    df_refacturation = df_refacturation.sort_values(
        ['Partenaire', 'TOTAL_A_REFACTURER'], 
        ascending=[True, False]
    )
    
    # Arrondir les montants
    for col in cols_numeriques + ['TOTAL_A_REFACTURER']:
        df_refacturation[col] = df_refacturation[col].round(2)
    
    # ==================== CRÉER LE FICHIER EXCEL ====================
    output = BytesIO()
    df_refacturation.to_excel(output, index=False, sheet_name='Refacturation', engine='openpyxl')
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb['Refacturation']
    
    # Mise en forme des en-têtes
    for cell in ws[1]:
        cell.fill = bleu_greenlog
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Mettre en jaune les cellules avec coûts > 0, en rouge le TOTAL
    for row_idx in range(2, ws.max_row + 1):
        # Alterner les couleurs de fond
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = vert_clair
        
        # Mettre en jaune les coûts > 0
        for col_idx in range(5, 10):  # DPD à Colis_Prive
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = jaune_alert
                cell.font = Font(bold=True)
        
        # Mettre en rouge le TOTAL
        total_cell = ws.cell(row=row_idx, column=10)
        if total_cell.value and total_cell.value > 0:
            total_cell.fill = rouge_alert
            total_cell.font = Font(bold=True, size=11, color='CC0000')
    
    # Ajuster largeur des colonnes
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Ajouter un onglet SYNTHÈSE PAR PARTENAIRE
    ws_synthese = wb.create_sheet('Synthèse Partenaire', 0)
    
    # Calculer synthèse
    synthese_data = []
    for partner in df_refacturation['Partenaire'].unique():
        df_partner = df_refacturation[df_refacturation['Partenaire'] == partner]
        
        row_data = {
            'Partenaire': partner,
            'Nb_Commandes': len(df_partner),
            'DPD': df_partner['DPD'].sum(),
            'Mondial_Relay': df_partner['Mondial_Relay'].sum(),
            'Colissimo': df_partner['Colissimo'].sum(),
            'Chronopost': df_partner['Chronopost'].sum(),
            'Colis_Prive': df_partner['Colis_Prive'].sum(),
            'TOTAL': df_partner['TOTAL_A_REFACTURER'].sum()
        }
        synthese_data.append(row_data)
    
    df_synthese = pd.DataFrame(synthese_data)
    df_synthese = df_synthese.sort_values('TOTAL', ascending=False)
    
    # Ajouter ligne de TOTAL GÉNÉRAL
    total_general = {
        'Partenaire': 'TOTAL GÉNÉRAL',
        'Nb_Commandes': df_synthese['Nb_Commandes'].sum(),
        'DPD': df_synthese['DPD'].sum(),
        'Mondial_Relay': df_synthese['Mondial_Relay'].sum(),
        'Colissimo': df_synthese['Colissimo'].sum(),
        'Chronopost': df_synthese['Chronopost'].sum(),
        'Colis_Prive': df_synthese['Colis_Prive'].sum(),
        'TOTAL': df_synthese['TOTAL'].sum()
    }
    df_synthese = pd.concat([df_synthese, pd.DataFrame([total_general])], ignore_index=True)
    
    # Écrire la synthèse
    ws_synthese['A1'] = "SYNTHÈSE REFACTURATION PAR PARTENAIRE"
    ws_synthese['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws_synthese['A1'].fill = bleu_greenlog
    ws_synthese.merge_cells('A1:H1')
    ws_synthese.row_dimensions[1].height = 30
    
    ws_synthese['A2'] = f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if partenaires_filter:
        ws_synthese['A3'] = f"Filtré : {', '.join(partenaires_filter)}"
    
    # En-têtes
    start_row = 5 if partenaires_filter else 4
    for col_idx, col_name in enumerate(df_synthese.columns, 1):
        cell = ws_synthese.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = bleu_greenlog
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Données
    for row_idx, row_data in enumerate(df_synthese.values, start_row + 1):
        is_total_row = row_data[0] == 'TOTAL GÉNÉRAL'
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_synthese.cell(row=row_idx, column=col_idx, value=value)
            
            if is_total_row:
                cell.fill = bleu_greenlog
                cell.font = Font(bold=True, color='FFFFFF', size=12)
            elif row_idx % 2 == 0:
                cell.fill = vert_clair
            
            # Colonne TOTAL en rouge
            if col_idx == 8 and not is_total_row and isinstance(value, (int, float)) and value > 0:
                cell.fill = rouge_alert
                cell.font = Font(bold=True, color='CC0000')
    
    # Ajuster colonnes synthèse
    for col_idx in range(1, ws_synthese.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws_synthese.max_row + 1):
            cell = ws_synthese.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws_synthese.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output
    
    # Couleurs GREENLOG
    bleu_greenlog = PatternFill(start_color='2D3E50', end_color='2D3E50', fill_type='solid')
    rouge_alert = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    vert_clair = PatternFill(start_color='E8F5F1', end_color='E8F5F1', fill_type='solid')
    
    # ==================== COLLECTER DONNÉES DPD ====================
    if 'dpd_data' in st.session_state and st.session_state.dpd_data:
        try:
            df_dpd = st.session_state.dpd_data['df_with_taxes'].copy()
            
            for _, row in df_dpd.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                # Filtrer par partenaire si nécessaire
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Tracking': tracking,
                        'Partenaire': partner,
                        'Date': row.get('Date de la commande', ''),
                        'Num_Commande_Origine': row.get("Numéro de commande d'origine", ''),
                        'Num_Commande_Partenaire': row.get('Numéro de commande partenaire', ''),
                    }
                
                all_data[tracking]['DPD_Montant_HT'] = row.get('Montant HT', 0)
                all_data[tracking]['DPD_Taxe_Fuel'] = row.get('Taxe fuel', 0)
                all_data[tracking]['DPD_Taxe_Securite'] = row.get('Taxe sécurité', 0)
                all_data[tracking]['DPD_Total_TTC'] = row.get('Montant total TTC', 0)
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte DPD: {str(e)}")
    
    # ==================== COLLECTER DONNÉES MONDIAL RELAY ====================
    if 'mondial_relay_data' in st.session_state and st.session_state.mondial_relay_data:
        try:
            df_mr = st.session_state.mondial_relay_data['df'].copy()
            
            for _, row in df_mr.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Tracking': tracking,
                        'Partenaire': partner,
                        'Date': row.get('Date de la commande', ''),
                        'Num_Commande_Origine': row.get("Numéro de commande d'origine", ''),
                        'Num_Commande_Partenaire': row.get('Numéro de commande partenaire', ''),
                    }
                
                all_data[tracking]['MondialRelay_Prix_Retour'] = row.get('Prix du retour', 0)
                all_data[tracking]['MondialRelay_Taxe_Fuel'] = row.get('Taxe fuel', 0)
                all_data[tracking]['MondialRelay_Total'] = row.get('Total avec taxe', 0)
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Mondial Relay: {str(e)}")
    
    # ==================== COLLECTER DONNÉES COLISSIMO ====================
    if 'colissimo_data' in st.session_state and st.session_state.colissimo_data:
        try:
            df_col = st.session_state.colissimo_data['df_matched'].copy()
            
            for _, row in df_col.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Tracking': tracking,
                        'Partenaire': partner,
                        'Date': row.get('Date de la commande', ''),
                        'Num_Commande_Origine': row.get("Numéro de commande d'origine", ''),
                        'Num_Commande_Partenaire': row.get('Numéro de commande partenaire', ''),
                    }
                
                all_data[tracking]['Colissimo_Tracking_8R'] = row.get('Tracking 8R', '')
                all_data[tracking]['Colissimo_Methode'] = row.get('Méthode', '')
                all_data[tracking]['Colissimo_Statut'] = row.get('Statut', '')
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Colissimo: {str(e)}")
    
    # ==================== COLLECTER DONNÉES CHRONOPOST ====================
    if 'chronopost_data' in st.session_state and st.session_state.chronopost_data:
        try:
            df_chrono = st.session_state.chronopost_data['df'].copy()
            
            for _, row in df_chrono.iterrows():
                tracking = str(row.get('Tracking', ''))
                partner = row.get('Partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Tracking': tracking,
                        'Partenaire': partner,
                        'Date': row.get('Date', ''),
                        'Num_Commande_Origine': row.get('Num_Commande_Origine', ''),
                        'Num_Commande_Partenaire': row.get('Num_Commande_Partenaire', ''),
                    }
                
                all_data[tracking]['Chronopost_Prix_Facture'] = row.get('Prix_Facture_HT', 0)
                all_data[tracking]['Chronopost_Prix_Theorique'] = row.get('Prix_Theorique_HT', 0)
                all_data[tracking]['Chronopost_Ecart_Prix'] = row.get('Difference_Prix', 0)
                all_data[tracking]['Chronopost_Poids_Facture'] = row.get('Poids_Chronopost', 0)
                all_data[tracking]['Chronopost_Poids_Logisticien'] = row.get('Poids_Logisticien', 0)
                # Chronopost_Ecart_Poids RETIRÉ sur demande utilisateur
            
            # Ajouter les surplus
            if 'df_surplus' in st.session_state.chronopost_data:
                df_surplus = st.session_state.chronopost_data['df_surplus'].copy()
                
                # Grouper les surplus par tracking
                surplus_by_tracking = df_surplus.groupby('Tracking').agg({
                    'Montant_Surplus': 'sum',
                    'Type_Surplus': lambda x: ', '.join(x.unique())
                }).to_dict('index')
                
                for tracking, surplus_data in surplus_by_tracking.items():
                    tracking_str = str(tracking)
                    if tracking_str in all_data:
                        all_data[tracking_str]['Chronopost_Surplus_Montant'] = surplus_data['Montant_Surplus']
                        all_data[tracking_str]['Chronopost_Surplus_Types'] = surplus_data['Type_Surplus']
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Chronopost: {str(e)}")
    
    # ==================== COLLECTER DONNÉES COLIS PRIVÉ ====================
    if 'colis_prive_data' in st.session_state and st.session_state.colis_prive_data:
        try:
            df_cp = st.session_state.colis_prive_data['df'].copy()
            
            for _, row in df_cp.iterrows():
                tracking = str(row.get('Numéro de tracking', ''))
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                if tracking not in all_data:
                    all_data[tracking] = {
                        'Tracking': tracking,
                        'Partenaire': partner,
                        'Date': row.get('Date de la commande', ''),
                        'Num_Commande_Origine': row.get("Numéro de commande d'origine", ''),
                        'Num_Commande_Partenaire': row.get('Numéro de commande partenaire', ''),
                    }
                
                # ColisPrive_Poids_Facture RETIRÉ sur demande utilisateur
                all_data[tracking]['ColisPrive_Majoration'] = row.get('Majoration service', 0)
                # ColisPrive_Code_Postal RETIRÉ sur demande utilisateur
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Colis Privé: {str(e)}")
    
    # ==================== COLLECTER DONNÉES RETOURS PRODUITS ====================
    if 'retours_data' in st.session_state and st.session_state.retours_data:
        try:
            df_retours = st.session_state.retours_data['df_grouped'].copy()
            
            for _, row in df_retours.iterrows():
                partner = row.get('Nom du partenaire', 'Non attribué')
                
                if partenaires_filter and partner not in partenaires_filter:
                    continue
                
                # Pour les retours, on n'a pas forcément de tracking unique
                # On va créer une entrée synthétique par partenaire
                key = f"RETOURS_{partner}"
                
                if key not in all_data:
                    all_data[key] = {
                        'Tracking': 'SYNTHESE_RETOURS',
                        'Partenaire': partner,
                        'Date': '',
                        'Num_Commande_Origine': '',
                        'Num_Commande_Partenaire': '',
                    }
                
                all_data[key]['Retours_Nb_Total'] = row.get('Nombre de retours', 0)
                all_data[key]['Retours_Nb_Avec_Motif'] = row.get('Avec motif', 0)
                all_data[key]['Retours_Nb_Sans_Motif'] = row.get('Sans motif', 0)
        
        except Exception as e:
            st.warning(f"⚠️ Erreur collecte Retours: {str(e)}")
    
    # ==================== CRÉER LE DATAFRAME CONSOLIDÉ ====================
    if not all_data:
        return None
    
    df_consolidated = pd.DataFrame.from_dict(all_data, orient='index')
    
    # Définir TOUTES les colonnes possibles (même si vides)
    all_possible_columns = {
        # Colonnes de base
        'Partenaire': None,
        'Tracking': None,
        'Date': None,
        'Num_Commande_Origine': None,
        'Num_Commande_Partenaire': None,
        
        # DPD
        'DPD_Montant_HT': 0,
        'DPD_Taxe_Fuel': 0,
        'DPD_Taxe_Securite': 0,
        'DPD_Total_TTC': 0,
        
        # Mondial Relay
        'MondialRelay_Prix_Retour': 0,
        'MondialRelay_Taxe_Fuel': 0,
        'MondialRelay_Total': 0,
        
        # Colissimo
        'Colissimo_Tracking_8R': None,
        'Colissimo_Methode': None,
        'Colissimo_Statut': None,
        
        # Chronopost
        'Chronopost_Prix_Facture': 0,
        'Chronopost_Prix_Theorique': 0,
        'Chronopost_Ecart_Prix': 0,
        'Chronopost_Poids_Facture': 0,
        'Chronopost_Poids_Logisticien': 0,
        'Chronopost_Surplus_Montant': 0,
        'Chronopost_Surplus_Types': None,
        
        # Colis Privé
        'ColisPrive_Majoration': 0,
        
        # Retours
        'Retours_Nb_Total': 0,
        'Retours_Nb_Avec_Motif': 0,
        'Retours_Nb_Sans_Motif': 0,
    }
    
    # Ajouter les colonnes manquantes avec leurs valeurs par défaut
    for col, default_value in all_possible_columns.items():
        if col not in df_consolidated.columns:
            df_consolidated[col] = default_value
    
    # Réorganiser les colonnes dans un ordre logique
    base_cols = ['Partenaire', 'Tracking', 'Date', 'Num_Commande_Origine', 'Num_Commande_Partenaire']
    
    dpd_cols = ['DPD_Montant_HT', 'DPD_Taxe_Fuel', 'DPD_Taxe_Securite', 'DPD_Total_TTC']
    mr_cols = ['MondialRelay_Prix_Retour', 'MondialRelay_Taxe_Fuel', 'MondialRelay_Total']
    col_cols = ['Colissimo_Tracking_8R', 'Colissimo_Methode', 'Colissimo_Statut']
    chrono_cols = ['Chronopost_Prix_Facture', 'Chronopost_Prix_Theorique', 'Chronopost_Ecart_Prix', 
                   'Chronopost_Poids_Facture', 'Chronopost_Poids_Logisticien', 
                   'Chronopost_Surplus_Montant', 'Chronopost_Surplus_Types']
    cp_cols = ['ColisPrive_Majoration']
    retours_cols = ['Retours_Nb_Total', 'Retours_Nb_Avec_Motif', 'Retours_Nb_Sans_Motif']
    
    ordered_cols = base_cols + dpd_cols + mr_cols + col_cols + chrono_cols + cp_cols + retours_cols
    
    df_consolidated = df_consolidated[ordered_cols]
    
    # Trier par partenaire puis par date
    df_consolidated = df_consolidated.sort_values(['Partenaire', 'Date'], ascending=[True, False])
    
    # Remplacer NaN/None par 0 pour les colonnes numériques
    numeric_cols = dpd_cols + mr_cols + chrono_cols + cp_cols + retours_cols
    for col in numeric_cols:
        if col in df_consolidated.columns:
            df_consolidated[col] = df_consolidated[col].fillna(0)
    
    # ==================== CRÉER LE FICHIER EXCEL ====================
    output = BytesIO()
    df_consolidated.to_excel(output, index=False, sheet_name='Export Consolidé', engine='openpyxl')
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb['Export Consolidé']
    
    # Mise en forme des en-têtes
    for cell in ws[1]:
        cell.fill = bleu_greenlog
        cell.font = Font(bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Identifier les colonnes de majorations/écarts pour mise en rouge
    alert_columns = []
    for idx, col_name in enumerate(df_consolidated.columns, 1):
        if any(keyword in col_name for keyword in ['Ecart', 'Majoration', 'Surplus_Montant']):
            alert_columns.append(idx)
    
    # Appliquer la mise en forme
    for row_idx in range(2, ws.max_row + 1):
        # Alterner les couleurs de fond
        if row_idx % 2 == 0:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = vert_clair
        
        # Mettre en rouge les cellules avec majorations/écarts > 0
        for col_idx in alert_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = rouge_alert
                cell.font = Font(bold=True, color='CC0000')
    
    # Ajuster largeur des colonnes
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Figer la première ligne
    ws.freeze_panes = 'A2'
    
    # Ajouter un onglet de synthèse par partenaire
    ws_synthese = wb.create_sheet('Synthèse par Partenaire', 0)
    
    # Grouper par partenaire
    synthese_data = []
    for partner in df_consolidated['Partenaire'].unique():
        df_partner = df_consolidated[df_consolidated['Partenaire'] == partner]
        
        row_data = {
            'Partenaire': partner,
            'Nb_Commandes': len(df_partner[df_partner['Tracking'] != 'SYNTHESE_RETOURS']),
        }
        
        # DPD
        if 'DPD_Total_TTC' in df_partner.columns:
            row_data['DPD_Total'] = df_partner['DPD_Total_TTC'].sum()
        
        # Mondial Relay
        if 'MondialRelay_Total' in df_partner.columns:
            row_data['MondialRelay_Total'] = df_partner['MondialRelay_Total'].sum()
        
        # Chronopost
        if 'Chronopost_Ecart_Prix' in df_partner.columns:
            row_data['Chronopost_Ecarts'] = df_partner['Chronopost_Ecart_Prix'].sum()
        if 'Chronopost_Surplus_Montant' in df_partner.columns:
            row_data['Chronopost_Surplus'] = df_partner['Chronopost_Surplus_Montant'].sum()
        
        # Colis Privé
        if 'ColisPrive_Majoration' in df_partner.columns:
            row_data['ColisPrive_Majorations'] = df_partner['ColisPrive_Majoration'].sum()
        
        # Retours
        if 'Retours_Nb_Total' in df_partner.columns:
            row_data['Retours_Total'] = df_partner['Retours_Nb_Total'].sum()
        
        # Total général
        row_data['TOTAL_GENERAL'] = (
            row_data.get('DPD_Total', 0) +
            row_data.get('MondialRelay_Total', 0) +
            row_data.get('Chronopost_Ecarts', 0) +
            row_data.get('Chronopost_Surplus', 0) +
            row_data.get('ColisPrive_Majorations', 0)
        )
        
        synthese_data.append(row_data)
    
    df_synthese = pd.DataFrame(synthese_data)
    df_synthese = df_synthese.sort_values('TOTAL_GENERAL', ascending=False)
    
    # Écrire la synthèse
    ws_synthese['A1'] = "SYNTHÈSE PAR PARTENAIRE"
    ws_synthese['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws_synthese['A1'].fill = bleu_greenlog
    ws_synthese.merge_cells('A1:I1')
    ws_synthese.row_dimensions[1].height = 30
    
    ws_synthese['A2'] = f"Date d'export : {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if partenaires_filter:
        ws_synthese['A3'] = f"Filtré : {', '.join(partenaires_filter)}"
    
    # En-têtes
    start_row = 5 if partenaires_filter else 4
    headers = list(df_synthese.columns)
    for col_idx, header in enumerate(headers, 1):
        cell = ws_synthese.cell(row=start_row, column=col_idx, value=header)
        cell.fill = bleu_greenlog
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Données
    for row_idx, row_data in enumerate(df_synthese.values, start_row + 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_synthese.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = vert_clair
            
            # Mettre en rouge les valeurs > 0 pour écarts/majorations
            if col_idx > 2 and isinstance(value, (int, float)) and value > 0:
                if 'Ecart' in headers[col_idx-1] or 'Majoration' in headers[col_idx-1] or 'Surplus' in headers[col_idx-1]:
                    cell.fill = rouge_alert
    
    # Ajuster colonnes synthèse
    for col_idx in range(1, ws_synthese.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws_synthese.max_row + 1):
            cell = ws_synthese.cell(row=row_idx, column=col_idx)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_synthese.column_dimensions[column_letter].width = adjusted_width
    
    # Sauvegarder
    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    
    return final_output

def run_export_interface():
    """Interface d'export global sur la page d'accueil"""
    
    st.markdown("---")
    st.subheader("💰 Export Refacturation Multi-Transporteurs")
    
    # Vérifier qu'il y a des données
    has_data = any([
        'dpd_data' in st.session_state and st.session_state.dpd_data,
        'mondial_relay_data' in st.session_state and st.session_state.mondial_relay_data,
        'colissimo_data' in st.session_state and st.session_state.colissimo_data,
        'chronopost_data' in st.session_state and st.session_state.chronopost_data,
        'colis_prive_data' in st.session_state and st.session_state.colis_prive_data,
        'retours_data' in st.session_state and st.session_state.retours_data
    ])
    
    if not has_data:
        st.info("ℹ️ Aucune donnée disponible. Analysez au moins un module transporteur pour pouvoir exporter.")
        return
    
    # Récupérer tous les partenaires uniques
    all_partners = set()
    
    for data_key, col_name in [
        ('dpd_data', 'Nom du partenaire'),
        ('mondial_relay_data', 'Nom du partenaire'),
        ('colissimo_data', 'Nom du partenaire'),
        ('colis_prive_data', 'Nom du partenaire'),
        ('retours_data', 'Nom du partenaire')
    ]:
        if data_key in st.session_state and st.session_state[data_key]:
            if data_key == 'dpd_data':
                df = st.session_state[data_key]['synthese']
            elif data_key == 'mondial_relay_data':
                df = st.session_state[data_key]['synthese']
            elif data_key == 'colissimo_data':
                df = st.session_state[data_key]['detail']
            elif data_key == 'colis_prive_data':
                df = st.session_state[data_key]['df']
            elif data_key == 'retours_data':
                df = st.session_state[data_key]['synthese']
            
            if col_name in df.columns:
                all_partners.update(df[col_name].dropna().unique())
    
    if 'chronopost_data' in st.session_state and st.session_state.chronopost_data:
        if 'Partenaire' in st.session_state.chronopost_data['df'].columns:
            all_partners.update(st.session_state.chronopost_data['df']['Partenaire'].dropna().unique())
    
    all_partners = sorted(list(all_partners))
    
    # Interface
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("""
        **Export simplifié pour REFACTURATION :**
        - 💰 **Focus : Coûts à refacturer uniquement**
        - 📋 Une ligne = une commande avec coûts
        - 💵 Colonnes : DPD | Mondial Relay | Colissimo | Chronopost | Colis Privé | **TOTAL**
        - 🟡 Les coûts > 0 sont surlignés en jaune
        - 🔴 Le TOTAL est en rouge pour visibilité
        - ✅ Synthèse par partenaire en 1er onglet
        """)
    
    with col2:
        st.info(f"**{len(all_partners)}** partenaire(s) détecté(s)")
    
    # Filtrage par partenaire
    st.markdown("### 🔍 Filtrage par Partenaire")
    
    filter_mode = st.radio(
        "Sélectionnez les partenaires à inclure dans l'export :",
        ["Tous les partenaires", "Sélection personnalisée"],
        horizontal=True
    )
    
    selected_partners = None
    
    if filter_mode == "Sélection personnalisée":
        if len(all_partners) > 0:
            selected_partners = st.multiselect(
                "Choisissez un ou plusieurs partenaires :",
                options=all_partners,
                default=None,
                help="Sélectionnez les partenaires dont vous voulez voir les données"
            )
            
            if selected_partners:
                st.success(f"✅ {len(selected_partners)} partenaire(s) sélectionné(s) : {', '.join(selected_partners)}")
            else:
                st.warning("⚠️ Aucun partenaire sélectionné - l'export sera vide")
        else:
            st.error("❌ Aucun partenaire disponible")
    else:
        st.info(f"ℹ️ Export de tous les partenaires ({len(all_partners)} partenaires)")
    
    # Bouton d'export
    st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        export_disabled = (filter_mode == "Sélection personnalisée" and not selected_partners)
        
        if st.button("📥 Générer Export Refacturation", type="primary", use_container_width=True, disabled=export_disabled):
            with st.spinner("Génération de l'export refacturation..."):
                try:
                    # Générer l'export
                    excel_file = create_consolidated_export(
                        partenaires_filter=selected_partners if filter_mode == "Sélection personnalisée" else None
                    )
                    
                    if excel_file is None:
                        st.error("❌ Aucune donnée à exporter avec les filtres sélectionnés")
                        return
                    
                    # Nom du fichier
                    filename_parts = ["Refacturation"]
                    if filter_mode == "Sélection personnalisée" and selected_partners:
                        if len(selected_partners) <= 3:
                            filename_parts.extend(selected_partners[:3])
                        else:
                            filename_parts.append(f"{len(selected_partners)}_partenaires")
                    filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
                    filename = "_".join(filename_parts) + ".xlsx"
                    
                    # Bouton de téléchargement
                    st.success("✅ Export généré avec succès !")
                    
                    st.download_button(
                        label=f"📥 Télécharger {filename}",
                        data=excel_file,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                
                except Exception as e:
                    st.error(f"❌ Erreur lors de la génération de l'export: {str(e)}")
                    import traceback
                    with st.expander("Détails de l'erreur"):
                        st.code(traceback.format_exc())
